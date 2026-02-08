from django.shortcuts import get_object_or_404
from django.shortcuts import render, redirect
from django.db.models import Sum, Max, Q, F, Value as V, CharField, DecimalField
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, Max
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect, FileResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
import os
import openpyxl
import qrcode
from io import BytesIO
from django.conf import settings
import uuid
import subprocess
from decimal import Decimal
import requests
import base64
import json
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
import pytz
from datetime import datetime, timedelta
from django.utils.timezone import make_aware, localtime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from xml.etree.ElementTree import Element, SubElement, tostring
from django.utils import timezone
from .forms import SearchForm
from .models import Invoices, Items, FotarahResponses
from .helpers import n2w, update_invoice
from PIL import Image as PILImage
from django.utils import timezone
from recipt_vouchers.models import Vouchers, Attachments
from drivers.models import Drivers, Treasury
from customers.models import Customers, Treasury as Tr
from drivers.views import update_driver_balance
from customers.views import update_customer_balance
from suppliers.views import update_vehicle
from mobile.models import Used_Vehicle

def index(request):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=6)
    invoices = Invoices.objects.filter(date__range=(start_date, end_date)).order_by('number')
    start_date = start_date.strftime("%Y-%m-%d")
    end_date = end_date.strftime("%Y-%m-%d")
    customers = Customers.objects.all()
    vehicles = Used_Vehicle.objects.filter(status="Available")
    form = SearchForm(request.GET)
    result = None
    options = "option0"
    amman_tz = pytz.timezone('Asia/Amman')
    type = "جميع الفواتير"
    if "search" in request.GET or "export" in request.GET:
        result = "ok"
        start_date = request.GET['start_date']
        end_date = request.GET['end_date']
        if form.is_valid():
            if request.GET['inlineRadioOptions'] == "option0":
                invoices = Invoices.objects.filter(date__range=(start_date, end_date)).order_by('number')
            elif request.GET['inlineRadioOptions'] == "option1":
                invoices = Invoices.objects.filter(Q(tamount__gt=F('driver_paid')) & Q(tamount__gt=F('customer_paid'))).exclude(type="كليك").order_by('number')
                options = "option1"
                type = "الفواتير الغير مسددة"
            elif request.GET['inlineRadioOptions'] == "option2":
                invoices = Invoices.objects.filter(Q(type="ذمم") & Q(date__range=(start_date, end_date))).order_by('number')
                options = "option2"
                type = "الفواتير الذمم"
            elif request.GET['inlineRadioOptions'] == "option3":
                invoices = Invoices.objects.filter(Q(type="نقدي") & Q(date__range=(start_date, end_date))).order_by('number')
                options = "option3"
                type = "الفواتير النقدية"
            elif request.GET['inlineRadioOptions'] == "option4":
                invoices = Invoices.objects.filter(Q(type="كليك") & Q(date__range=(start_date, end_date))).order_by('number')
                options = "option4"
                type = "الفواتير الكليك"
            elif request.GET['inlineRadioOptions'] == "option5":
                invoices = Invoices.objects.filter(exported__isnull=True).order_by('number')
                options = "option5"
                type = "الفواتير الغير مرحلة"
            query = form.cleaned_data['query']
            #print("okok", len(query))
            if len(query) > 0:
                try:
                    invoices = invoices.filter(Q(number=query) | Q(car__icontains=query))
                except:
                    invoices = invoices.filter(Q(driver__name__icontains=query) | Q(require_id__name__icontains=query) | Q(car__icontains=query) | Q(notes__icontains=query))
    number_of_invoices = invoices.count()
    total_invoiced = invoices.aggregate(Sum('tamount'))['tamount__sum'] or 0
    adjusted_invoices = []
    for invoice in invoices:
        invoice.created_at = localtime(invoice.created_at, amman_tz)
        invoice.adjusted_created_at = invoice.created_at
        time_diff = invoice.created_at.utcoffset().total_seconds() / 3600
        if invoice.created_at:
            invoice.adjusted_created_at += timedelta(hours=time_diff)
        adjusted_invoices.append(invoice)
    if "export" in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Invoices_Report.xlsx"'
        source_excel_path = os.path.join(settings.BASE_DIR, 'excel_templates', 'invoices.xlsx')
        output_excel_path = os.path.join(settings.BASE_DIR, 'excel_templates', 'Invoices_Report.xlsx')
        # Load the existing workbook
        wb = load_workbook(source_excel_path)
        ws = wb.active
        # Set the title in cell A1
        ws['A1'] = f"بيان {type} للفترة من {start_date} وحتى {end_date}"
        ws['D5003'] = total_invoiced
        # Write data dynamically and track the last written row
        last_row = 2  # Start after header row
        for idx, item in enumerate(adjusted_invoices, start=3):
            ws.cell(row=idx, column=1, value=item.number)
            ws.cell(row=idx, column=2, value=item.date)
            ws.cell(row=idx, column=3, value=item.type)
            ws.cell(row=idx, column=4, value=item.require_id.name)
            ws.cell(row=idx, column=5, value=item.tamount)
            ws.cell(row=idx, column=6, value=item.car)
            last_row = idx  # Update last row index
        # Hide or delete rows from last_row + 1 to 5001
        for row in range(last_row + 1, 5002):  # 5001 inclusive
            ws.row_dimensions[row].hidden = True  # Hides rows instead of deleting
        # Save the modified workbook
        wb.save(output_excel_path)
        # Return Excel file as response
        with open(output_excel_path, 'rb') as f:
            response.write(f.read())
        return response
    else:
        paginator = Paginator(adjusted_invoices, 10)  # 10 invoices per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context = {
            'active_tab': 'invoices',
            'start_date': start_date,
            'end_date': end_date,
            'result': result,
            'options': options,
            'form': form,
            'invoices': adjusted_invoices,
            'page_obj': page_obj,
            'number_of_invoices': number_of_invoices,
            'total_invoiced': total_invoiced,
            'customers': customers,
            'vehicles': vehicles
        }
        if request.method == "POST":
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('invoices_index')
            else:
                messages.success(request, "There was an error for username or password...")
                return redirect('invoices_index')
        else:
            username = request.user.username
            if 'jobbox' in username:
                return redirect('pmissions_index')
            else:
                return render(request, "invoices/index.html", context)

def invoice_details(request, number):
    invoice = get_object_or_404(Invoices, number=number)
    items = Items.objects.filter(number=invoice.pk)
    adjusted_items = []
    for item in items:
        item.tamount = item.udinar + (item.ufils / 100)
        adjusted_items.append(item)
    invoice_id = invoice.id
    number = invoice.number
    customer_id = invoice.require_id.id
    car = invoice.car
    date = invoice.date
    system = invoice.type
    tamount = invoice.tamount
    samount = invoice.samount
    notes = invoice.notes
    driver = invoice.driver
    name = Customers.objects.get(id=customer_id).name
    phone = Customers.objects.get(id=customer_id).phone
    try:
        mail = Customers.objects.get(id=customer_id).mail
    except Customers.DoesNotExist:
        mail = None
    context = {
        'invoice_id': invoice_id,
        'number': number,
        'items': adjusted_items,
        'name': name,
        'car': car,
        'date': date,
        'phone': phone,
        'mail': mail,
        'system': system,
        'notes': notes,
        'tamount': tamount,
        'samount': samount,
        'driver': driver,
        'active_tab': 'invoices',
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('invoices_index')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('invoices_index')
    else:
        return render(request, "invoices/details.html", context)

@login_required
def edit_invoice(request, id):
    invoice = get_object_or_404(Invoices, id=id)
    if request.method == 'POST':
        vehicle_number = request.POST.get('vehicle')
        vehicle = get_object_or_404(Used_Vehicle, vehicle_number=vehicle_number)
        old_vehicle = invoice.vehicle
        invoice.number = request.POST.get('number')
        invoice.date = request.POST.get('date')
        invoice.type = request.POST.get('system')
        invoice.vehicle = vehicle
        invoice.exported = request.POST.get('exported')
        invoice.notes = request.POST.get('notes')
        tinvoice = Invoices.objects.exclude(id=id)
        if not tinvoice:
            invoice.save()
            update_customer_balance(invoice.require_id.id)
            update_driver_balance(invoice.driver.name)
            if old_vehicle.vehicle_number != invoice.vehicle.vehicle_number:
                update_vehicle(old_vehicle.vehicle_number)
                update_vehicle(invoice.vehicle.vehicle_number)
        else:
            queryset = tinvoice.filter(number=request.POST.get('number'))
            if queryset.exists():
                record = queryset.first()  # This will get the first object that matches the query
            else:
                record = None
            if record:
                messages.success(request, "Repeated Invoice Number...")
            else:
                invoice.save()
                update_customer_balance(invoice.require_id.id)
                update_driver_balance(invoice.driver.name)
                if old_vehicle.vehicle_number != invoice.vehicle.vehicle_number:
                    update_vehicle(old_vehicle.vehicle_number)
                    update_vehicle(invoice.vehicle.vehicle_number)
    return redirect('invoices_index')

# Configuration
FOTARAH_API_URL = "https://fotarah.markaziaapis.com/api/fotarah-invoices/Send"
FOTARAH_TOKEN = "b87d7ece-0638-4eaa-8116-30f5370c332f"

@login_required
def send_invoice(request, number):
    try:
        invoice = Invoices.objects.get(number=number)
    except Invoices.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    try:
        items = Items.objects.filter(number__number=number)
    except Items.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    voucher = Vouchers.objects.filter(invoice__id=invoice.id).first()
    invoice_type = "Cash"
    if invoice.type == "ذمم":
        invoice_type = "Receivable"
    # Prepare invoice data for Fotarah API
    invoice_data = {
        "SystemName": "SMT",
        "InvoiceNo": str(invoice.number),
        "InvoiceName": "SMT",
        "Entity": "AMMAN",
        "Company": "SMT",
        "InvoiceType": "Sales",
        "TaxType": "Income",
        "SalesType": "Sales",
        "PaymentMethod": invoice_type,
        "BusinessArea": "Amman",
        "InvoiceID": invoice.number,
        "InvoiceYear": invoice.date.year,
        "InvoiceDate": invoice.date.isoformat(),
        "InvoiceTime": "00:00:00",
        "CustomerNo": invoice.require_id.id if invoice.require_id else 0,
        "CustomerName": invoice.require_id.name if invoice.require_id else "",
        "CustomerIdentityType": "",
        "CustomerIdentityNo": "",
        "CustomerPostalZone": "",
        "CustomerCity": "",
        "CustomerTelephone": "",
        "Currency": "JOD",
        "Total": float(invoice.tamount) if invoice.tamount else 0,
        "DiscountAmount": 0,
        "TotalNet": float(invoice.tamount) if invoice.tamount else 0,
        "TaxAmount": 0,
        "GrandTotal": float(invoice.tamount) if invoice.tamount else 0,
        "Note": "",
        "OriginalInvoiceNo": "",
        "invoiceLines": [
            {
                "LineID": idx + 1,
                "LineSeq": idx + 1,
                "LineDesc": "نقل مركبة للعميل",
                "LineQTY": item.qty,
                "LinePrice": float(item.udinar) + (float(item.ufils) / 100) if item.adinar else 0,
                "LineAmount": float(item.adinar) + (float(item.afils) / 100) if item.adinar else 0,
                "LineDiscountPercentage": 0,
                "LineDiscountAmount": 0,
                "LineNetAmount": float(item.adinar) + (float(item.afils) / 100) if item.adinar else 0,
                "LineTaxPercentage": 0,
                "LineTaxAmount": 0,
                "LineTotalAmount": float(item.adinar) + (float(item.afils) / 100) if item.adinar else 0,
                "LineTaxable": True
            }
            for idx, item in enumerate(items)
        ]
    }

    # Send to Fotarah API
    # Initialize QR code variables
    qr_temp_path = None
    qr_data = None
    qr_id = None
    duplicate_error = False
    use_stored_qr = False

    # Check if we have a previously successful response
    previous_response = FotarahResponses.objects.filter(
        invoice=invoice,
        success=True
    ).order_by('-created_at').first()

    # Send to Fotarah API only if not previously sent successfully
    if not invoice.fotarah_success:
        headers = {
            "accept": "*/*",
            "AuthorizationKey": FOTARAH_TOKEN,
            "Content-Type": "application/json"
        }
        try:
            response = requests.post(FOTARAH_API_URL, headers=headers, json=invoice_data)
            response_data = response.json()
            # Check for duplicate key error
            if "Cannot insert duplicate key in object" in str(response_data):
                duplicate_error = True
                # If we have a previous successful response with QR code
                if previous_response and previous_response.qr_code_base64:
                    qr_data = previous_response.qr_code_base64
                    qr_id = previous_response.fotarah_invoice_id
                    use_stored_qr = True
                else:
                    # No stored QR code available
                    pass
            # Save API response regardless of success
            fotarah_response = FotarahResponses.objects.create(
                invoice=invoice,
                success=response_data.get('success', False) and not duplicate_error,
                message=response_data.get('message', ''),
                fotarah_invoice_id=response_data.get('fotarahInvoiceID'),
                qr_code_base64=response_data.get('fotarahQRCodeBase64'),
                qr_code_binary=response_data.get('fotarahQRBinary'),
                response_data=response_data
            )

            # Update invoice status
            invoice.fotarah_sent = True
            invoice.fotarah_success = response_data.get('success', False) and not duplicate_error
            invoice.save()

            if not duplicate_error and not response_data.get('success', False):
                return HttpResponse("Failed to send invoice to Fotarah: " +
                                    response_data.get('message', 'Unknown error'),
                                    status=status.HTTP_400_BAD_REQUEST)

            # Get QR data from current response if not duplicate error
            if not duplicate_error:
                qr_data = response_data.get('fotarahQRCodeBase64')
                qr_id = response_data.get('fotarahInvoiceID')
        except Exception as e:
            # Save failed attempt
            FotarahResponses.objects.create(
                invoice=invoice,
                success=False,
                message=str(e),
                response_data={"error": str(e)}
            )
            invoice.fotarah_sent = True
            invoice.fotarah_success = False
            invoice.save()
            # Continue with export even if API fails
    else:
        # Use previous successful response
        if previous_response:
            qr_data = previous_response.qr_code_base64
            qr_id = previous_response.fotarah_invoice_id
    # Generate QR code image if we have data
    if qr_data:
        try:
            '''
            # Decode base64 string
            decoded_qr_data = base64.b64decode(qr_data)
            try:
                decoded_qr_data = decoded_qr_data.decode("utf-8")
            except UnicodeDecodeError:
                pass  # Use as bytes if not UTF-8
            '''
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=3,
                border=1,
            )
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            # Save QR image to temp path
            qr_temp_path = os.path.join(settings.MEDIA_ROOT, f'temp_qr_{number}.png')
            qr_img.save(qr_temp_path)
        except Exception as e:
            print(f"Failed to generate QR code: {str(e)}")
            qr_temp_path = None
    # Prepare Excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Invoice Number {number}.xlsx"'
    source_excel_path = os.path.join(settings.BASE_DIR, 'excel_templates', 'tax_invoice.xlsx')
    output_excel_path = f"{os.path.splitext(source_excel_path)[0]} Number {number}.xlsx"
    # Load the existing workbook
    wb = load_workbook(source_excel_path)
    ws = wb.active
    # Add voucher signature if exists
    if voucher:
        img = OpenpyxlImage(voucher.receipt_to_signature.path)
        ws.add_image(img, 'B33')
    # Add QR code to Excel only if we have one
    if qr_temp_path and os.path.exists(qr_temp_path):
        try:
            qr_excel_img = OpenpyxlImage(qr_temp_path)
            # Position at cell H5 (adjust as needed)
            ws.add_image(qr_excel_img, 'H5')
            # Add note about duplicate if applicable
            if duplicate_error and use_stored_qr:
                ws['H36'] = "QR code from previous submission"
        except Exception as e:
            print(f"Failed to add QR code to Excel: {str(e)}")
    # Fill invoice data
    name = invoice.require_id.name if invoice.require_id else ""
    sub_customer = ""
    if invoice.require_id and invoice.require_id.notes:
        if "الشرك" in invoice.require_id.notes and "المركزي" in invoice.require_id.notes and "تجار" in invoice.require_id.notes:
            sub_customer = name
            name = invoice.require_id.notes
    ws['A6'] = number
    ws['E7'] = invoice.type
    ws['H8'] = qr_id
    ws['G9'] = invoice.date
    ws['B11'] = name
    ws['A13'] = invoice.car
    ws['B31'] = invoice.samount
    ws['H32'] = invoice.dinar
    ws['I32'] = invoice.fils

    if voucher:
        ws['E34'] = voucher.recipient_to_name
    ws['E35'] = sub_customer
    ws['E36'] = invoice.driver.name if invoice.driver else ""
    ws['E38'] = invoice.notes if invoice.notes else ""

    for idx, item in enumerate(items, start=19):
        ws.cell(row=idx, column=1, value=item.udinar)
        ws.cell(row=idx, column=2, value=item.ufils)
        ws.cell(row=idx, column=3, value=item.qty)
        ws.cell(row=idx, column=4, value=item.description)
        ws.cell(row=idx, column=8, value=item.adinar)
        ws.cell(row=idx, column=9, value=item.afils)

    # Save Excel file
    wb.save(output_excel_path)

    try:
        output_pdf_path = os.path.splitext(output_excel_path)[0] + '.pdf'

        # Kill any existing LibreOffice processes
        subprocess.run(["pkill", "-f", "soffice.bin"], check=False)

        # Convert to PDF
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf", output_excel_path, '--outdir',
             os.path.dirname(output_pdf_path)],
            check=True
        )

        if os.path.exists(output_pdf_path):
            invoice.exported = "Done"
            invoice.save()

            # Clean up temporary files
            if qr_temp_path and os.path.exists(qr_temp_path):
                os.remove(qr_temp_path)

            with open(output_pdf_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{os.path.basename(output_pdf_path)}"'
            return response
        else:
            raise Http404("PDF file not found")

    except Exception as e:
        print("Error:", e)
        return HttpResponse(f"Error during PDF conversion: {str(e)}",
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@login_required
def export_invoice(request, number):
    try:
        invoice = Invoices.objects.get(number=number)
    except Invoices.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    try:
        items = Items.objects.filter(number__number=number)
    except Items.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    voucher = Vouchers.objects.filter(invoice__id=invoice.id).first()

    # Send to Fotarah API
    # Initialize QR code variables
    qr_temp_path = None
    qr_data = None
    qr_id = None
    # Check if we have a previously successful response
    previous_response = FotarahResponses.objects.filter(
        invoice=invoice,
        success=True
    ).order_by('-created_at').first()
    # Use previous successful response
    if previous_response:
        qr_data = previous_response.qr_code_base64
        qr_id = previous_response.fotarah_invoice_id
    # Generate QR code image if we have data
    if qr_data:
        try:
            '''
            # Decode base64 string
            decoded_qr_data = base64.b64decode(qr_data)
            try:
                decoded_qr_data = decoded_qr_data.decode("utf-8")
            except UnicodeDecodeError:
                pass  # Use as bytes if not UTF-8
            '''
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=3,
                border=1,
            )
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            # Save QR image to temp path
            qr_temp_path = os.path.join(settings.MEDIA_ROOT, f'temp_qr_{number}.png')
            qr_img.save(qr_temp_path)
        except Exception as e:
            print(f"Failed to generate QR code: {str(e)}")
            qr_temp_path = None
    # Prepare Excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Invoice Number {number}.xlsx"'
    source_excel_path = os.path.join(settings.BASE_DIR, 'excel_templates', 'tax_invoice.xlsx')
    output_excel_path = f"{os.path.splitext(source_excel_path)[0]} Number {number}.xlsx"
    # Load the existing workbook
    wb = load_workbook(source_excel_path)
    ws = wb.active
    # Add voucher signature if exists
    if voucher:
        img = OpenpyxlImage(voucher.receipt_to_signature.path)
        ws.add_image(img, 'B33')

    # Add QR code to Excel only if we have one
    if qr_temp_path and os.path.exists(qr_temp_path):
        try:
            qr_excel_img = OpenpyxlImage(qr_temp_path)
            # Position at cell H5 (adjust as needed)
            ws.add_image(qr_excel_img, 'H5')
            # Add note about duplicate if applicable
            if duplicate_error and use_stored_qr:
                ws['H36'] = "QR code from previous submission"
        except Exception as e:
            print(f"Failed to add QR code to Excel: {str(e)}")

    # Fill invoice data
    name = invoice.require_id.name if invoice.require_id else ""
    sub_customer = ""
    if invoice.require_id and invoice.require_id.notes:
        if "الشرك" in invoice.require_id.notes and "المركزي" in invoice.require_id.notes and "تجار" in invoice.require_id.notes:
            sub_customer = name
            name = invoice.require_id.notes

    ws['A6'] = number
    ws['E7'] = invoice.type
    if qr_id:
        ws['H8'] = qr_id
    ws['G9'] = invoice.date
    ws['B11'] = name
    ws['A13'] = invoice.car
    ws['B31'] = invoice.samount
    ws['H32'] = invoice.dinar
    ws['I32'] = invoice.fils
    if voucher:
        ws['E34'] = voucher.recipient_to_name
    ws['E35'] = sub_customer
    ws['E36'] = invoice.driver.name if invoice.driver else ""
    ws['E38'] = invoice.notes if invoice.notes else ""

    for idx, item in enumerate(items, start=19):
        ws.cell(row=idx, column=1, value=item.udinar)
        ws.cell(row=idx, column=2, value=item.ufils)
        ws.cell(row=idx, column=3, value=item.qty)
        ws.cell(row=idx, column=4, value=item.description)
        ws.cell(row=idx, column=8, value=item.adinar)
        ws.cell(row=idx, column=9, value=item.afils)

    # Save Excel file
    wb.save(output_excel_path)

    try:
        output_pdf_path = os.path.splitext(output_excel_path)[0] + '.pdf'

        # Kill any existing LibreOffice processes
        subprocess.run(["pkill", "-f", "soffice.bin"], check=False)

        # Convert to PDF
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf", output_excel_path, '--outdir',
             os.path.dirname(output_pdf_path)],
            check=True
        )

        if os.path.exists(output_pdf_path):
            #invoice.exported = "Done"
            #invoice.save()

            # Clean up temporary files
            if qr_temp_path and os.path.exists(qr_temp_path):
                os.remove(qr_temp_path)

            with open(output_pdf_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{os.path.basename(output_pdf_path)}"'
            return response
        else:
            raise Http404("PDF file not found")

    except Exception as e:
        print("Error:", e)
        return HttpResponse(f"Error during PDF conversion: {str(e)}",
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@login_required
def add_invitem(request):
    if request.method == 'POST':
        number = request.POST['number']
        driver = request.POST['driver']
        invoice = get_object_or_404(Invoices, id=number)
        description = request.POST['description']
        qty = int(request.POST['qty'])
        # Parsing and rounding the amount
        pre_amount = round(float(request.POST['amount']), 2)
        # Extracting udinar (integer part) and ufils (fractional part as fils)
        udinar = int(pre_amount)  # Dinar part
        ufils = int(round((pre_amount - udinar) * 100))  # Fils part (fractional part multiplied by 100)
        # Calculate the total amount for adinar and afils
        total_amount = round(pre_amount * qty, 2)
        adinar = int(total_amount)  # Dinar part for the total amount
        afils = int(round((total_amount - adinar) * 100))  # Fils part for the total amount
        # Add the new item to the database
        new_item = Items(number=invoice, description=f"\u202B{description}\u202C", qty=qty, udinar=udinar, ufils=ufils, adinar=adinar, afils=afils)
        new_item.save()
        update_invoice(number)
        update_customer_balance(invoice.require_id.id)
        update_driver_balance(driver)
        update_vehicle(invoice.vehicle.vehicle_number)
        related_url = reverse('invoice_details', kwargs={'number': invoice.number})
        return HttpResponseRedirect(related_url)

@login_required
def edit_invitem(request, id):
    item = get_object_or_404(Items, id=id)
    number = item.number
    if request.method == 'POST':
        # Update description and quantity
        item.description = f"\u202B{request.POST['description']}\u202C"
        item.qty = float(request.POST['qty'])
        # Parsing and rounding the amount
        pre_amount = round(float(request.POST['amount']), 2)
        # Extract dinar and fils from the pre_amount
        udinar = int(pre_amount)  # Dinar part
        ufils = int(round((pre_amount - udinar) * 100))  # Fils part
        # Update the item's udinar and ufils
        item.udinar = udinar
        item.ufils = ufils
        # Calculate the total amount for adinar and afils
        amount = round(pre_amount * item.qty, 2)
        adinar = int(amount)  # Dinar part for total amount
        afils = int(round((amount - adinar) * 100))  # Fils part for total amount
        # Update the item's adinar and afils
        item.adinar = adinar
        item.afils = afils
        # Save the updated item
        item.save()
    update_invoice(number.id)
    update_customer_balance(number.require_id.id)
    update_driver_balance(number.driver)
    update_vehicle(number.vehicle.vehicle_number)
    related_url = reverse('invoice_details', kwargs={'number': number.number})
    return HttpResponseRedirect(related_url)

@login_required
def delete_invitem(request, id):
    item = get_object_or_404(Items, id=id)
    number = item.number
    if request.method == 'POST':
        item.delete()
    update_invoice(number.id)
    update_customer_balance(number.require_id.id)
    update_driver_balance(number.driver)
    update_vehicle(number.vehicle.vehicle_number)
    related_url = reverse('invoice_details', kwargs={'number': number.number})
    return HttpResponseRedirect(related_url)


def generate_xml(invoice):
    root = Element('Invoice', {
        "xmlns": "urn:oasis:names:specification:ubl:schema:xsd:Invoice-2",
        "xmlns:cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
        "xmlns:cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
        "xmlns:ext": "urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2"
    })
    # Invoice Header
    SubElement(root, 'cbc:ID').text = str(invoice.number)
    SubElement(root, 'cbc:UUID').text = str(uuid.uuid4())
    SubElement(root, 'cbc:IssueDate').text = invoice.date.strftime('%Y-%m-%d')
    SubElement(root, 'cbc:InvoiceTypeCode').text = invoice.invoice_type_code
    SubElement(root, 'cbc:DocumentCurrencyCode').text = invoice.currency_code
    SubElement(root, 'cbc:TaxCurrencyCode').text = invoice.currency_code
    # Seller Information
    supplier = SubElement(root, 'cac:AccountingSupplierParty')
    party = SubElement(supplier, 'cac:Party')
    party_tax_scheme = SubElement(party, 'cac:PartyTaxScheme')
    SubElement(party_tax_scheme, 'cbc:CompanyID').text = "12345678"
    tax_scheme = SubElement(party_tax_scheme, 'cac:TaxScheme')
    SubElement(tax_scheme, 'cbc:ID').text = "VAT"  # VAT should be inside <cbc:ID>
    party_legal_entity = SubElement(party, 'cac:PartyLegalEntity')
    SubElement(party_legal_entity, 'cbc:RegistrationName').text = "Seller Name"
    # Buyer Information
    customer = SubElement(root, 'cac:AccountingCustomerParty')
    party = SubElement(customer, 'cac:Party')
    party_identification = SubElement(party, 'cac:PartyIdentification')
    SubElement(party_identification, 'cbc:ID', schemeID="TN").text = "33445544"
    # Postal Address
    postal_address = SubElement(party, 'cac:PostalAddress')
    SubElement(postal_address, 'cbc:PostalZone').text = "33554"
    country = SubElement(postal_address, 'cac:Country')
    SubElement(country, 'cbc:IdentificationCode').text = "JO"
    # Tax Scheme
    party_tax_scheme = SubElement(party, 'cac:PartyTaxScheme')
    tax_scheme = SubElement(party_tax_scheme, 'cac:TaxScheme')
    SubElement(tax_scheme, 'cbc:ID').text = "VAT"
    # Party Legal Entity
    party_legal_entity = SubElement(party, 'cac:PartyLegalEntity')
    SubElement(party_legal_entity, 'cbc:RegistrationName').text = "Ahmad Mohammad"
    # Accounting Contact
    accounting_contact = SubElement(customer, 'cac:AccountingContact')
    SubElement(accounting_contact, 'cbc:Telephone').text = "777888999"
    # **Payment & Tax Information (Must be before InvoiceLine)**
    tax_total = SubElement(root, 'cac:TaxTotal')
    SubElement(tax_total, 'cbc:TaxAmount', currencyID="JOD").text = "10.00"
    legal_total = SubElement(root, 'cac:LegalMonetaryTotal')
    SubElement(legal_total, 'cbc:PayableAmount', currencyID="JOD").text = str(invoice.tamount)
    # **Invoice Lines (Should come last)**
    for item in invoice.invoice_items.all():
        invoice_line = SubElement(root, 'cac:InvoiceLine')
        SubElement(invoice_line, 'cbc:ID').text = str(item.id)
        SubElement(invoice_line, 'cbc:InvoicedQuantity', unitCode="PCE").text = str(item.qty)
        line_extension = (Decimal(item.udinar) + Decimal(item.ufils) / 100) * Decimal(item.qty) - item.discount
        SubElement(invoice_line, 'cbc:LineExtensionAmount', currencyID="JOD").text = str(line_extension)
        item_element = SubElement(invoice_line, 'cac:Item')
        SubElement(item_element, 'cbc:Name').text = item.description
        price = SubElement(invoice_line, 'cac:Price')
        SubElement(price, 'cbc:PriceAmount', currencyID="JOD").text = str(item.udinar + item.ufils / 100)
        allowance_charge = SubElement(price, 'cac:AllowanceCharge')
        SubElement(allowance_charge, 'cbc:ChargeIndicator').text = "false"
        SubElement(allowance_charge, 'cbc:AllowanceChargeReason').text = "DISCOUNT"
        SubElement(allowance_charge, 'cbc:Amount', currencyID="JOD").text = str(item.discount)
    # Convert to XML string
    xml_str = tostring(root, encoding='utf-8', method='xml').decode('utf-8')
    return xml_str

def encode_xml_to_base64(xml_string):
    xml_bytes = xml_string.encode('utf-8')
    base64_bytes = base64.b64encode(xml_bytes)
    return base64_bytes.decode('utf-8')

'''
def send_invoice_to_jofotara(invoice):
    # Generate the XML
    xml_content = generate_xml(invoice)
    # Encode the XML in Base64
    encoded_xml = base64.b64encode(xml_content.encode('utf-8')).decode('utf-8')
    # Prepare the JSON payload
    payload = {
        "invoice": encoded_xml
    }
    # Send the request to JoFotara API
    headers = {
        "Client-id": "f5600953-7047-4fb7-b36e-e948a9f8bffd",
        "Secret-Key": "Gj5nS9wyYHRadaVffz5VKB4v4wlVWyPhcJvrTD4NHtM4dET01wPHg+WSAZ9Yw+nVb1tlBQgZ7/bx7NfP07ge92/C7jHwgZCZ+j154LpRcfuBWuDJfJ598ylvrmJ1tOTKCS7/AVlxeRuLlmnnJTptoEzEg8+NtMd0jTDGvwjN/Vnf0+ebq5Emikar3Mjke7J2Z0sK9bXT2gByv7YEZ8YLLD7IgntQgU2pYG5zMmsdN5TzcWaCfEywSMYKrM4cI69QIiRYk7/MQi9LoEORKdtllA==",
        "Content-Type": "application/json"
    }
    response = requests.post("https://backend.jofotara.gov.jo/core/invoices/", json=payload, headers=headers)
    # Handle the response
    # **Debugging: Print full response**
    print("Response Status Code:", response.status_code)
    print("Response Headers:", response.headers)
    print("Response Content:", response.text)
    return response.text
'''

def send_to_jofotara(request, id):
    invoice = get_object_or_404(Invoices, id=id)
    #invoice_xml = generate_xml(invoice)  # Generate XML first
    #encoded_xml = encode_xml_to_base64(invoice_xml)  # Encode XML to Base64
    #return render(request, 'invoices/send_to_jofotara.html', {'response': encoded_xml})

    # Generate XML
    invoice_xml = generate_xml(invoice)

    # Encode the XML
    encoded_xml = encode_xml_to_base64(invoice_xml)

    # API endpoint
    url = "https://backend.jofotara.gov.jo/core/invoices/"

    # Headers
    headers = {
        "Client-id": "f5600953-7047-4fb7-b36e-e948a9f8bffd",
        "Secret-Key": "Gj5nS9wyYHRadaVffz5VKB4v4wlVWyPhcJvrTD4NHtM4dET01wPHg+WSAZ9Yw+nVb1tlBQgZ7/bx7NfP07ge92/C7jHwgZCZ+j154LpRcfuBWuDJfJ598ylvrmJ1tOTKCS7/AVlxeRuLlmnnJTptoEzEg8+NtMd0jTDGvwjN/Vnf0+ebq5Emikar3Mjke7J2Z0sK9bXT2gByv7YEZ8YLLD7IgntQgU2pYG5zMmsdN5TzcWaCfEywSMYKrM4cI69QIiRYk7/MQi9LoEORKdtllA==",
        "Content-Type": "application/json",
        "Cookie": "stickounet=4fdb7136e666916d0e373058e9e5c44e|7480c8b0e4ce7933ee164081a50488f1"
    }

    # JSON Payload
    payload = json.dumps({
        "invoice": encoded_xml
    })

    # Send request
    response = requests.post(url, headers=headers, data=payload)

    # Print Response
    print(response.status_code, response.text)
