from django.shortcuts import get_object_or_404
from django.shortcuts import render, redirect
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, Max, Q, F, Value as V, CharField, DecimalField
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect, FileResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
import os
import openpyxl
from django.conf import settings
import subprocess
import zipfile
from django.core.files.storage import FileSystemStorage
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
import pytz
from django.utils.timezone import localtime
from datetime import datetime, timedelta, date
from django.contrib.auth.models import User
from .forms import SearchForm
from .models import Supplier, Treasury, MainItems, Invoices, Items, Papers
from mobile.models import Used_Vehicle, Auto_Oil, Fuel_Filling
from invoices.models import Invoices as Inv
from invoices.helpers import n2w

def index(request):
    new_treasury = Treasury.objects.aggregate(Max('number'))['number__max']
    if new_treasury is None:
        # If no records exist in the table, set the invoice number to 1
        new_treasury = 1
    else:
        # Add 1 to the maximum invoice number
        new_treasury += 1
    new_number = Invoices.objects.aggregate(Max('number'))['number__max']
    if new_number is None:
        # If no records exist in the table, set the invoice number to 1
        new_number = 1
    else:
        # Add 1 to the maximum invoice number
        new_number += 1
    suppliers = Supplier.objects.all().order_by('-balance')
    form = SearchForm(request.GET)
    result = None
    if "search" in request.GET:
        result = "ok"
        if form.is_valid():
            query = form.cleaned_data['query']
            try:
                query_value = float(query)
                suppliers = suppliers.filter(balance__gte=query_value)
            except:
                suppliers = suppliers.filter(Q(name__icontains=query) | Q(phone__icontains=query) | Q(mail__icontains=query) | Q(notes__icontains=query))
    #amman_tz = pytz.timezone('Asia/Amman')
    #adjusted_suppliers = []
    #for supplier in suppliers:
        #supplier.created_at = localtime(supplier.created_at, amman_tz)
        #supplier.adjusted_created_at = supplier.created_at
        #time_diff = supplier.adjusted_created_at.utcoffset().total_seconds() / 3600
        #supplier.adjusted_created_at += timedelta(hours=time_diff)
        #adjusted_suppliers.append(supplier)
    number_of_suppliers = suppliers.count()
    total_balances = suppliers.aggregate(Sum('balance'))['balance__sum'] or 0
    context = {
        'active_tab': 'suppliers',
        'new_treasury': new_treasury,
        'new_number': new_number,
        'result': result,
        'form': form,
        'suppliers': suppliers,
        'number_of_suppliers': number_of_suppliers,
        'total_balances': total_balances
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('suppliers_index')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('suppliers_index')
    else:
        username = request.user.username
        if 'jobbox' in username:
            return redirect('pmissions_index')
        else:
            return render(request, "suppliers/index.html", context)

def get_invoices(request):
    supplier_id = request.GET.get('supplier_id')
    invoices = Invoices.objects.filter(Q(supplier__id=supplier_id) & Q(tamount__gt=F('paid')))
    invoice_list = list(invoices.values('number'))
    return JsonResponse({'invoices': invoice_list})

def get_invoice_details(request):
    invoice_number = request.GET.get('invoice_number')
    invoice = Invoices.objects.filter(number=invoice_number).first()
    if invoice:
        amount_due = invoice.tamount - invoice.paid
        return JsonResponse({'amount_due': amount_due})
    return JsonResponse({'amount_due': 0})

def supplier_details(request, id):
    supplier = get_object_or_404(Supplier, id=id)
    papers = Papers.objects.filter(invoice__supplier__id=id)
    invoices = Invoices.objects.filter(supplier_id__id=id).annotate(
        account=V('Created invoice', output_field=CharField()),
        number_annotated=F('number'),
        amount_annotated=F('tamount'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'created_at_annotated'
    )
    treasury = Treasury.objects.filter(supplier__id=id).annotate(
        account=V('Paid to treasury', output_field=CharField()),
        number_annotated=F('number'),
        amount_annotated=F('amount'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'created_at_annotated'
    )
    combined_entries = invoices.union(treasury).order_by('created_at_annotated')
    context = {
        'active_tab': 'suppliers',
        'id': id,
        'supplier': supplier,
        'papers': papers,
        'combined_entries': combined_entries,
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('suppliers_index')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('suppliers_index')
    else:
        return render(request, "suppliers/details.html", context)

def vehicle_details(request, id):
    vehicle = Used_Vehicle.objects.filter(id=id).first()
    vehicle_number = vehicle.vehicle_number
    vehicle_meter = max(vehicle.oil_meter, vehicle.fuel_meter)
    context = {
        'active_tab': 'vehicles',
        'id': id,
        'vehicle_number': vehicle_number,
        'vehicle_meter': vehicle_meter
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('vehicle_details')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('vehicle_details')
    else:
        return render(request, "suppliers/vehicle_details.html", context)

def export_invoice(request, id):
    return redirect('suppliers_index')

def vehicle_balance(request, id):
    vehicle = Used_Vehicle.objects.filter(id=id).first()
    vehicle_number = vehicle.vehicle_number
    context = {
        'active_tab': 'vehicles',
        'id': id,
        'vehicle_number': vehicle_number,
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('vehicle_balance')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('vehicle_balance')
    else:
        return render(request, "suppliers/vehicle_balance.html", context)

def supplier_balance(request, id):
    supplier = get_object_or_404(Supplier, id=id)
    supplier_name = supplier.name
    invoices = Invoices.objects.filter(Q(supplier__id=id) & Q(type="ذمم"))
    treasury = Treasury.objects.filter(supplier__id=id)
    total_invoiced = invoices.aggregate(Sum('tamount'))['tamount__sum'] or 0
    total_treasury = treasury.aggregate(Sum('amount'))['amount__sum'] or 0
    required_amount = total_invoiced - total_treasury
    context = {
        'id': id,
        'supplier_name': supplier_name,
        'required_amount': required_amount,
        'invoices': invoices,
        'treasury': treasury,
        'active_tab': 'suppliers',
        'total_invoiced': total_invoiced,
        'total_treasury': total_treasury
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('supplier_balance')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('supplier_balance')
    else:
        return render(request, "suppliers/balance.html", context)

@login_required
def add_supplier(request):
    if request.method == 'POST':
        name = request.POST.get('supplier_name')
        mail = request.POST.get('supplier_mail')
        notes = request.POST.get('supplier_notes')
        phone = request.POST.get('supplier_phone')
        logo = request.FILES.get('supplier_logo')
        queryset = Supplier.objects.filter(Q(name=name) | Q(phone=phone))
        if mail:
            queryset = Supplier.objects.filter(Q(name=name) | Q(phone=phone) | Q(mail=mail))
        if queryset.exists():
            record = queryset.first()  # This will get the first object that matches the query
        else:
            record = None
        if record:
            messages.success(request, "Repeated Data...")
        else:
            if logo:
                fs = FileSystemStorage()
                logo_name = fs.save(logo.name, logo)
                # Save to the database
                Supplier.objects.create(name=name, mail=mail, phone=phone, notes=notes, logo=logo_name, balance=0.0)
            else:
                Supplier.objects.create(name=name, mail=mail, phone=phone, notes=notes, balance=0.0)
    return redirect('suppliers_index')

@login_required
def edit_supplier(request, id):
    supplier = get_object_or_404(Supplier, id=id)
    if request.method == 'POST':
        supplier.name = request.POST.get('name')
        supplier.phone = request.POST.get('phone')
        supplier.mail = request.POST.get('mail')
        supplier.notes = request.POST.get('notes')
        supplier_logo = request.FILES.get('logo')
        if supplier_logo:
            fs = FileSystemStorage()
            logo_name = fs.save(supplier_logo.name, supplier_logo)
            supplier.logo = logo_name
        tsupplier = Supplier.objects.exclude(id=id)
        if not tsupplier:
            supplier.save()
        else:
            queryset = tsupplier.filter(Q(name=request.POST.get('name')) | Q(phone=request.POST.get('phone')))
            if request.POST.get('mail'):
                queryset = tsupplier.filter(Q(name=request.POST.get('name')) | Q(phone=request.POST.get('phone')) | Q(mail=request.POST.get('mail')))
            if queryset.exists():
                record = queryset.first()  # This will get the first object that matches the query
            else:
                record = None
            if record:
                messages.success(request, "Repeated Data...")
            else:
                supplier.save()
    return redirect('suppliers_index')

@login_required
def add_attachment(request):
    if request.method == 'POST':
        supplier_id = request.POST.get('supplier_id')
        supplier = get_object_or_404(Supplier, id=supplier_id)
        description = request.POST.get('description')
        attachment = request.FILES.get('attachment')
        fs = FileSystemStorage()
        attachment_name = fs.save(attachment.name, attachment)
        # Save to the database
        Papers.objects.create(supplier=supplier, description=description, paper=attachment_name)
        related_url = reverse('supplier_details', kwargs={'id': supplier_id})
        return HttpResponseRedirect(related_url)

@login_required
def edit_attachment(request, id):
    paper = get_object_or_404(Papers, id=id)
    if request.method == 'POST':
        supplier_id = paper.supplier.id
        paper.description = request.POST.get('description')
        supplier_paper = request.FILES.get('paper')
        if supplier_paper:
            fs = FileSystemStorage()
            paper_name = fs.save(supplier_paper.name, supplier_paper)
            paper.paper = paper_name
        paper.save()
        related_url = reverse('supplier_details', kwargs={'id': supplier_id})
        return HttpResponseRedirect(related_url)

@login_required
def delete_attachment(request, id):
    paper = get_object_or_404(Papers, id=id)
    if request.method == 'POST':
        supplier_id = paper.supplier.id
        paper.delete()
        related_url = reverse('supplier_details', kwargs={'id': supplier_id})
        return HttpResponseRedirect(related_url)

@login_required
def add_treasury(request):
    if request.method == 'POST':
        supplier_id = request.POST['supplier_id']
        supplier = get_object_or_404(Supplier, id=supplier_id)
        invoice_number = request.POST['invoice_number']
        invoice = get_object_or_404(Invoices, number=invoice_number)
        number = request.POST['number']
        amount = request.POST['amount']
        payment = request.POST['method']
        # Add the new item to the database
        new_treasury = Treasury(supplier=supplier, invoice=invoice, number=number, amount=amount, payment=payment)
        new_treasury.save()
        update_supplier_balance(supplier_id)
        update_invoice(invoice.id)
    return redirect('suppliers_index')

@login_required
def edit_treasury(request, id):
    if request.method == 'POST':
        supplier_id = request.POST['supplier_id']
        item = get_object_or_404(Treasury, id=id)
        item.amount = request.POST['edit-amount']
        item.payment = request.POST['method']
        item.save()
        update_supplier_balance(supplier_id)
        update_invoice(item.invoice.id)
        related_url = reverse('supplier_balance', kwargs={'id': supplier_id})
        return HttpResponseRedirect(related_url)

@login_required
def delete_treasury(request, id):
    item = get_object_or_404(Treasury, id=id)
    invoice_id = item.invoice.id
    if request.method == 'POST':
        supplier_id = request.POST['supplier_id']
        item.delete()
        update_supplier_balance(supplier_id)
        update_invoice(invoice_id)
        related_url = reverse('supplier_balance', kwargs={'id': supplier_id})
        return HttpResponseRedirect(related_url)

@login_required
def invoice_details(request, supplier_id, invoice_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    invoice = get_object_or_404(Invoices, id=invoice_id)
    items = Items.objects.filter(invoice__id=invoice_id)
    main_items = MainItems.objects.all()
    vehicles = Used_Vehicle.objects.filter(status="Available")
    context = {
        'active_tab': 'suppliers',
        'supplier': supplier,
        'invoice': invoice,
        'items': items,
        'main_items': main_items,
        'vehicles': vehicles
    }
    return render(request, "suppliers/invoice.html", context)

@login_required
def add_invoice(request):
    if request.method == 'POST':
        supplier_id = request.POST['supplier_id']
        supplier = get_object_or_404(Supplier, id=supplier_id)
        number = request.POST['number']
        date = request.POST['date']
        type = request.POST['system']
        notes = request.POST['notes']
        system_user = request.user.username
        new_invoice = Invoices(number=number, date=date, type=type, supplier=supplier, notes=notes, created_by=system_user)
        new_invoice.save()
        related_url = reverse('invoice_details', kwargs={'supplier_id': supplier_id, 'invoice_id': new_invoice.id})
        return HttpResponseRedirect(related_url)

@login_required
def add_item(request):
    if request.method == 'POST':
        invoice_id = request.POST['invoice']
        invoice = get_object_or_404(Invoices, id=invoice_id)
        supplier_id = request.POST['supplier']
        vehicle_number = request.POST['vehicle']
        vehicle = get_object_or_404(Used_Vehicle, vehicle_number=vehicle_number)
        item = request.POST['item']
        description = request.POST['description']
        qty = int(request.POST['qty'])
        amount = float(request.POST['amount'])
        tamount = round(qty * amount, 2)
        system_user = request.user.username
        # Add the new item to the database
        new_item = Items(invoice=invoice, vehicle=vehicle, item=item, description=description, qty=qty, amount=amount, tamount=tamount, created_by=system_user)
        new_item.save()
        update_invoice(invoice_id)
        update_supplier_balance(supplier_id)
        update_vehicle(vehicle.vehicle_number)
        related_url = reverse('invoice_details', kwargs={'supplier_id': supplier_id, 'invoice_id': invoice_id})
        return HttpResponseRedirect(related_url)

@login_required
def edit_item(request, id):
    item = get_object_or_404(Items, id=id)
    if request.method == 'POST':
        supplier_id = request.POST['supplier']
        vehicle_number = item.vehicle.vehicle_number
        invoice_id = item.invoice.id
        item.item = request.POST['item']
        item.description = request.POST['description']
        item.qty = int(request.POST['qty'])
        item.amount = float(request.POST['amount'])
        item.tamount = round(item.qty * item.amount, 2)
        item.created_by = request.user.username
        item.save()
        update_invoice(invoice_id)
        update_supplier_balance(supplier_id)
        update_vehicle(vehicle_number)
        related_url = reverse('invoice_details', kwargs={'supplier_id': supplier_id, 'invoice_id': invoice_id})
        return HttpResponseRedirect(related_url)

@login_required
def delete_item(request, id):
    item = get_object_or_404(Items, id=id)
    if request.method == 'POST':
        supplier_id = request.POST['supplier']
        vehicle_number = item.vehicle.vehicle_number
        invoice_id = item.invoice.id
        item.delete()
        update_invoice(invoice_id)
        update_supplier_balance(supplier_id)
        update_vehicle(vehicle_number)
        related_url = reverse('invoice_details', kwargs={'supplier_id': supplier_id, 'invoice_id': invoice_id})
        return HttpResponseRedirect(related_url)

@login_required
def pocurement(request):
    invoices = Invoices.objects.all().order_by('number')
    form = SearchForm(request.GET)
    result = None
    start_date = None
    end_date = None
    if "search" in request.GET:
        result = "ok"
        start_date = request.GET['start_date']
        end_date = request.GET['end_date']
        invoices = invoices.filter(Q(date__range=(start_date, end_date)) | Q(tamount__gt=F('paid')))
        if form.is_valid():
            query = form.cleaned_data['query']
            try:
                query_value = float(query)
                invoices = invoices.filter(
                    Q(number__icontains=query) |
                    Q(tamount__gte=query_value))
            except:
                invoices = invoices.filter(
                    Q(type__icontains=query) |
                    Q(supplier__name__icontains=query)
                    )
    number_of_invoices = invoices.count()
    total_invoiced = invoices.aggregate(Sum('tamount'))['tamount__sum'] or 0
    context = {
        'active_tab': 'procurement',
        'start_date': start_date,
        'end_date': end_date,
        'result': result,
        'form': form,
        'invoices': invoices,
        'number_of_invoices': number_of_invoices,
        'total_invoiced': total_invoiced,
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('pocurement')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('pocurement')
    else:
        username = request.user.username
        if 'jobbox' in username:
            return redirect('pmissions_index')
        else:
            return render(request, "suppliers/procurement.html", context)

@login_required
def vehicles(request):
    vehicles = Used_Vehicle.objects.all()
    form = SearchForm(request.GET)
    result = None
    if "search" in request.GET:
        result = "ok"
        if form.is_valid():
            query = form.cleaned_data['query']
            vehicles = vehicles.filter(
                Q(vehicle_number__icontains=query) |
                Q(type__icontains=query)
                )
    number_of_vehicles = vehicles.count()
    context = {
        'active_tab': 'vehicles',
        'result': result,
        'form': form,
        'vehicles': vehicles,
        'number_of_vehicles': number_of_vehicles,
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('vehicles')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('vehicles')
    else:
        username = request.user.username
        if 'jobbox' in username:
            return redirect('pmissions_index')
        else:
            return render(request, "suppliers/vehicles.html", context)

@login_required
def edit_invoice(request, id):
    invoice = get_object_or_404(Invoices, id=id)
    if request.method == 'POST':
        system_user = request.user.username
        invoice.number = request.POST.get('number')
        invoice.date = request.POST.get('date')
        invoice.type = request.POST.get('system')
        invoice.notes = request.POST.get('notes')
        invoice.created_by = system_user
        tinvoice = Invoices.objects.exclude(id=id)
        if not tinvoice:
            invoice.save()
        else:
            queryset = tinvoice.filter(number=request.POST.get('number'))
            if queryset.exists():
                record = queryset.first()  # This will get the first object that matches the query
            else:
                record = None
            if record:
                messages.success(request, "Repeated Invoice Number...")
            else:
                invoice.save()
    update_supplier_balance(invoice.supplier.id)
    return redirect('procurement')

@login_required
def delete_invoice(request, id):
    invoice = get_object_or_404(Invoices, id=id)
    if request.method == 'POST':
        invoice.delete()
        update_supplier_balance(invoice.supplier.id)
        return redirect('procurement')

@login_required
def add_vehicle(request):
    if request.method == 'POST':
        vehicle_number = request.POST.get('vehicle_number').strip()
        vehicle_type = request.POST.get('vehicle_type')
        service_date = request.POST.get('service_date')
        status = request.POST.get('status')
        oil = request.POST.get('oil')
        fuel = request.POST.get('fuel')
        air = request.POST.get('air')
        notes = request.POST.get('notes')
        system_user = request.user.username
        vehicle_image = request.FILES.get('vehicle_image')
        if vehicle_image:
            fs = FileSystemStorage()
            vehicle_image_name = fs.save(vehicle_image.name, vehicle_image)
            # Save to the database
            Used_Vehicle.objects.create(vehicle_number=vehicle_number, type=vehicle_type, date=service_date, balance=0.0, status=status, default_oil_diff=oil, default_fuel_filter_diff=fuel, default_air_filter_diff=air, notes=notes, created_by=system_user, image=vehicle_image_name)
        else:
            Used_Vehicle.objects.create(vehicle_number=vehicle_number, type=vehicle_type, date=service_date, balance=0.0, status=status, default_oil_diff=oil, default_fuel_filter_diff=fuel, default_air_filter_diff=air, notes=notes, created_by=system_user)
    return redirect('vehicles')

@login_required
def edit_vehicle(request, id):
    vehicle = get_object_or_404(Used_Vehicle, id=id)
    if request.method == 'POST':
        vehicle.vehicle_number = request.POST.get('vehicle_number').strip()
        vehicle.type = request.POST.get('vehicle_type')
        vehicle.date = request.POST.get('service_date')
        vehicle.status = request.POST.get('status')
        vehicle.default_oil_diff = request.POST.get('oil')
        vehicle.default_fuel_filter_diff = request.POST.get('fuel')
        vehicle.default_air_filter_diff = request.POST.get('air')
        vehicle.notes = request.POST.get('notes')
        vehicle_image = request.FILES.get('vehicle_image')
        if vehicle_image:
            fs = FileSystemStorage()
            vehicle_image_name = fs.save(vehicle_image.name, vehicle_image)
            vehicle.image = vehicle_image_name
        tvehicle = Used_Vehicle.objects.exclude(id=id)
        if not tvehicle:
            vehicle.save()
        else:
            queryset = tvehicle.filter(vehicle_number=request.POST.get('vehicle_number'))
            if queryset.exists():
                record = queryset.first()  # This will get the first object that matches the query
            else:
                record = None
            if record:
                messages.success(request, "Repeated Vehicle Number...")
            else:
                vehicle.save()
    return redirect('vehicles')

@login_required
def export_fuel(request, id, date_from, date_to):
    if isinstance(date_to, str):
        # Convert the string to a datetime object if necessary
        date_to = datetime.strptime(date_to, '%Y-%m-%d')  # Adjust format as needed
    date_to += timedelta(days=1)
    vehicle = get_object_or_404(Used_Vehicle, id=id)
    vehicle_number = vehicle.vehicle_number
    oils = Auto_Oil.objects.filter(Q(vehicle__id=id) & Q(created_at__range=(date_from, date_to)))
    fuels = Fuel_Filling.objects.filter(Q(vehicle__id=id) & Q(created_at__range=(date_from, date_to)))
    source_excel_path = os.path.join(settings.BASE_DIR, 'excel_templates', 'Vehicle Oil Diesel Sheet.xlsx')
    output_excel_path = f"{os.path.splitext(source_excel_path)[0]}_{vehicle_number}.xlsx"
    # Load the existing workbook (source file)
    wb = openpyxl.load_workbook(source_excel_path)
    ws = wb.active
    amman_tz = pytz.timezone('Asia/Amman')
    cell = ws.cell(row=2, column=1)
    cell.value = "كشف حركة غيار الزيت الخاص بالمركبة " + "No. " + vehicle_number
    for row in range(2, 11289):
        ws.row_dimensions[row].hidden = False
    idx = start_row = 4
    for idx, item in enumerate(oils, start=start_row):
        air_filter = "✔" if item.air_filter else "✖"
        diesel_filter = "✔" if item.diesel_filter else "✖"
        # Convert created_at to Amman time zone
        created_at = localtime(item.created_at, amman_tz)
        ws.cell(row=idx, column=1, value=item.driver.name)
        ws.cell(row=idx, column=2, value=item.vehicle_meter)
        ws.cell(row=idx, column=3, value=item.maintenance_center)
        ws.cell(row=idx, column=4, value=diesel_filter)
        ws.cell(row=idx, column=5, value=air_filter)
        ws.cell(row=idx, column=6, value=created_at.replace(tzinfo=None))
    last_data_row = idx + 1
    if last_data_row >= 5:
        for value in range(last_data_row, 3002):
            ws.row_dimensions[value].hidden = True
    else:
        for value in range(2, 3002):
            ws.row_dimensions[value].hidden = True
    cell = ws.cell(row=3002, column=1)
    cell.value = "كشف حركة الديزل الخاص بالمركبة " + "No. " + vehicle_number
    idx = start_row = 3004
    for idx, item in enumerate(fuels, start=start_row):
        created_at = localtime(item.created_at, amman_tz)
        ws.cell(row=idx, column=1, value=item.driver.name)
        ws.cell(row=idx, column=2, value=item.vehicle_meter)
        ws.cell(row=idx, column=3, value=item.litres)
        ws.cell(row=idx, column=4, value=item.amount)
        ws.cell(row=idx, column=6, value=created_at.replace(tzinfo=None))
    last_data_row = idx + 1
    if last_data_row >= 3005:
        for value in range(last_data_row, 11289):
            ws.row_dimensions[value].hidden = True
    else:
        for value in range(3002, 11289):
            ws.row_dimensions[value].hidden = True
    wb.save(output_excel_path)
    try:
        output_pdf_path = os.path.splitext(output_excel_path)[0] + '.pdf'
        subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf", output_excel_path, '--outdir', os.path.dirname(output_pdf_path)], check=True)
        zip_path = f"{os.path.splitext(output_excel_path)[0]}.zip"
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(output_excel_path, os.path.basename(output_excel_path))
            zipf.write(output_pdf_path, os.path.basename(output_pdf_path))
        if os.path.exists(zip_path):
            with open(zip_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename="Vehicle_Report_{vehicle_number}.zip"'
                return response
        else:
            raise Http404("ZIP file not found")
    except Exception as e:
        print("Error:", e)
        raise Http404("An error occurred while generating the files.")

@login_required
def export_profit(request, id, date_from, date_to):
    vehicle = get_object_or_404(Used_Vehicle, id=id)
    vehicle_number = vehicle.vehicle_number
    revenue = Inv.objects.filter(Q(vehicle__id=id) & Q(date__range=(date_from, date_to))).annotate(
        account=V('إيرادات', output_field=CharField()),
        number_annotated=F('number'),
        amount_annotated=F('tamount'),
        customer_annotated=F('require_id__name'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'customer_annotated',
        'created_at_annotated'
    )
    expenses = Items.objects.filter(Q(vehicle__id=id) & Q(invoice__date__range=(date_from, date_to))).annotate(
        account=V('مصروفات', output_field=CharField()),
        number_annotated=F('invoice__number'),
        amount_annotated=F('tamount'),
        customer_annotated=F('invoice__supplier__name'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'customer_annotated',
        'created_at_annotated'
    )
    combined_entries = revenue.union(expenses).order_by('created_at_annotated')
    source_excel_path = os.path.join(settings.BASE_DIR, 'excel_templates', 'Vehicle Account Sheet.xlsx')
    output_excel_path = f"{os.path.splitext(source_excel_path)[0]}_{vehicle_number}.xlsx"
    wb = openpyxl.load_workbook(source_excel_path)
    ws = wb.active
    cell = ws.cell(row=2, column=1)
    cell.value = "كشف حركة الحساب الخاص بالمركبة " + "No. " + vehicle_number
    white_fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in range(4, 33428):
        ws.row_dimensions[row].hidden = False
    for row in ws.iter_rows(min_row=4, max_row=33428, max_col=5):
        for cell in row:
            cell.value = None
            cell.fill = white_fill
    idx = start_row = 4
    yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    amman_tz = pytz.timezone('Asia/Amman')
    for idx, item in enumerate(combined_entries, start=start_row):
        item['created_at_annotated'] = localtime(item['created_at_annotated'], amman_tz)
        ws.cell(row=idx, column=1, value=item['account'])
        ws.cell(row=idx, column=2, value=item['number_annotated'])
        ws.cell(row=idx, column=3, value=item['amount_annotated'])
        ws.cell(row=idx, column=4, value=item['customer_annotated'])
        ws.cell(row=idx, column=5, value=item['created_at_annotated'].replace(tzinfo=None))
        if item['account'] == "مصروفات":
            for col in range(1, 6):
                ws.cell(row=idx, column=col).fill = yellow_fill
    last_data_row = idx + 1
    for value in range(last_data_row, 33428):
        ws.row_dimensions[value].hidden = True
    wb.save(output_excel_path)
    try:
        output_pdf_path = f"{os.path.splitext(output_excel_path)[0]}.pdf"
        subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf", output_excel_path, '--outdir', os.path.dirname(output_pdf_path)], check=True)
        zip_path = f"{os.path.splitext(output_excel_path)[0]}.zip"
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(output_excel_path, os.path.basename(output_excel_path))
            zipf.write(output_pdf_path, os.path.basename(output_pdf_path))
        if os.path.exists(zip_path):
            with open(zip_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename="Vehicle_Report_{vehicle_number}.zip"'
                return response
        else:
            raise Http404("ZIP file not found")
    except Exception as e:
        print("Error:", e)
        raise Http404("An error occurred while generating the files.")

@login_required
def export_treasury(request, id, date_from, date_to):
    supplier = get_object_or_404(Supplier, id=id)
    supplier_name = supplier.name
    invoices = Invoices.objects.filter(Q(supplier__id=id) & Q(type="ذمم") & (Q(date__range=(date_from, date_to)) | Q(tamount__gt=F('paid')))).annotate(
        account=V('فاتورة', output_field=CharField()),
        number_annotated=F('number'),
        amount_annotated=F('tamount'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'created_at_annotated'
    )
    treasury = Treasury.objects.filter(Q(supplier__id=id) & (Q(invoice__date__range=(date_from, date_to)) | Q(invoice__tamount__gt=F('invoice__paid')))).annotate(
        account=V('سداد', output_field=CharField()),
        number_annotated=F('invoice__number'),
        amount_annotated=F('amount'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'created_at_annotated'
    )
    combined_entries = invoices.union(treasury).order_by('created_at_annotated')
    source_excel_path = os.path.join(settings.BASE_DIR, 'excel_templates', 'Supplier Account Sheet.xlsx')
    output_excel_path = f"{os.path.splitext(source_excel_path)[0]}_{supplier_name}.xlsx"
    # Load the existing workbook (source file)
    wb = openpyxl.load_workbook(source_excel_path)
    ws = wb.active
    cell = ws.cell(row=2, column=1)
    cell.value = "كشف حركة الحساب الخاص بالمورد/ " + supplier_name
    white_fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in range(4, 33428):
        ws.row_dimensions[row].hidden = False
    for row in ws.iter_rows(min_row=4, max_row=33428, max_col=4):
        for cell in row:
            cell.value = None
            cell.fill = white_fill
    idx = start_row = 4
    yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    amman_tz = pytz.timezone('Asia/Amman')
    for idx, item in enumerate(combined_entries, start=start_row):
        item['created_at_annotated'] = localtime(item['created_at_annotated'], amman_tz)
        ws.cell(row=idx, column=1, value=item['account'])
        ws.cell(row=idx, column=2, value=item['number_annotated'])
        ws.cell(row=idx, column=3, value=item['amount_annotated'])
        ws.cell(row=idx, column=4, value=item['created_at_annotated'].replace(tzinfo=None))
        if item['account'] == "سداد":
            for col in range(1, 5):  # Assuming you have data in columns 1 to 4
                ws.cell(row=idx, column=col).fill = yellow_fill
    last_data_row = idx + 1
    for value in range(last_data_row, 33428):
        ws.row_dimensions[value].hidden = True
    wb.save(output_excel_path)
    try:
        output_pdf_path = f"{os.path.splitext(output_excel_path)[0]}.pdf"
        subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf", output_excel_path, '--outdir', os.path.dirname(output_pdf_path)], check=True)
        zip_path = f"{os.path.splitext(output_excel_path)[0]}.zip"
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(output_excel_path, os.path.basename(output_excel_path))
            zipf.write(output_pdf_path, os.path.basename(output_pdf_path))
        if os.path.exists(zip_path):
            with open(zip_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename="Supplier_Report_{id}.zip"'
                return response
        else:
            raise Http404("ZIP file not found")
    except Exception as e:
        print("Error:", e)
        raise Http404("An error occurred while generating the files.")

@login_required
def view_att(request, id):
    result = "ok"
    invoices = Invoices.objects.filter(id=id)
    number_of_invoices = invoices.count()
    context = {
        'active_tab': 'procurement',
        'result': result,
        'invoices': invoices,
        'number_of_invoices': number_of_invoices
    }
    return render(request, "suppliers/procurement.html", context)

@login_required
def add_att(request, id):
    if request.method == 'POST':
        invoice = get_object_or_404(Invoices, id=id)
        description = request.POST.get('description')
        attachment = request.FILES.get('attachment')
        fs = FileSystemStorage()
        attachment_name = fs.save(attachment.name, attachment)
        # Save to the database
        Papers.objects.create(invoice=invoice, description=description, paper=attachment_name)
        related_url = reverse('view_att', kwargs={'id': id})
        return HttpResponseRedirect(related_url)

@login_required
def delete_att(request, id):
    if request.method == 'POST':
        paper = Papers.objects.filter(invoice__id=id).last()
        if paper:
            paper.delete()
        related_url = reverse('view_att', kwargs={'id': id})
        return HttpResponseRedirect(related_url)

def sfetch_attachments(request):
    invoice_id = request.GET.get('id')
    attachments = Papers.objects.filter(invoice__id=invoice_id)
    attachments_list = [
        {
            'paper_url': request.build_absolute_uri(attachment.paper.url),
            'paper_description': attachment.description
        }
        for attachment in attachments
    ]
    return JsonResponse({'attachments': attachments_list})

def update_supplier_balance(supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    total_invoiced = Invoices.objects.filter(Q(supplier__id=supplier_id) & Q(type="ذمم")).aggregate(Sum('tamount'))['tamount__sum'] or 0
    total_treasury = Treasury.objects.filter(supplier__id=supplier_id).aggregate(Sum('amount'))['amount__sum'] or 0
    supplier.balance = total_invoiced - total_treasury
    supplier.save()

def update_vehicle(vehicle_number):
    vehicle = get_object_or_404(Used_Vehicle, vehicle_number=vehicle_number)
    vehicle_debit = Items.objects.filter(vehicle__vehicle_number=vehicle_number).aggregate(Sum('tamount'))['tamount__sum'] or 0
    vehicle_credit = Inv.objects.filter(vehicle__vehicle_number=vehicle_number).aggregate(Sum('tamount'))['tamount__sum'] or 0
    vehicle.income = vehicle_credit
    vehicle.expenses = vehicle_debit
    vehicle.balance = float(vehicle_credit) - float(vehicle_debit)
    vehicle.save()

def bfilter_invoices(request, id):
    date_from = datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d')
    date_to = datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d')
    invoices = Inv.objects.filter(Q(vehicle__id=id) & Q(date__range=(date_from, date_to))).order_by('created_at')
    amman_tz = pytz.timezone('Asia/Amman')
    invoices_data = []
    for invoice in invoices:
        invoice.created_at = localtime(invoice.created_at, amman_tz)
        invoices_data.append({
            'number': invoice.number,
            'require_id__name': invoice.require_id.name,
            'tamount': invoice.tamount,
            'created_at': invoice.created_at
        })
    procurement = Items.objects.filter(Q(vehicle__id=id) & Q(invoice__date__range=(date_from, date_to))).order_by('created_at')
    procurement_data = []
    for pro in procurement:
        pro.created_at = localtime(pro.created_at, amman_tz)
        procurement_data.append({
            'number': pro.invoice.number,
            'invoice__supplier__name': pro.invoice.supplier.name,
            'tamount': pro.tamount,
            'created_at': pro.created_at
        })
    total_revenue = invoices.aggregate(Sum('tamount'))['tamount__sum'] or 0
    total_expenses = procurement.aggregate(Sum('tamount'))['tamount__sum'] or 0
    net_profit = total_revenue - total_expenses
    return JsonResponse({'invoices': invoices_data, 'procurement': procurement_data, 'total_revenue': total_revenue, 'total_expenses': total_expenses, 'net_profit': net_profit}, safe=False)

def vfilter_invoices(request, id):
    date_from = datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d')
    date_to = datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d')
    date_to += timedelta(days=1)
    oils = Auto_Oil.objects.filter(Q(vehicle__id=id) & Q(created_at__range=(date_from, date_to))).order_by('created_at')
    amman_tz = pytz.timezone('Asia/Amman')
    oils_data = []
    for oil in oils:
        oil.created_at = localtime(oil.created_at, amman_tz)
        oils_data.append({
            'driver__name': oil.driver.name,
            'vehicle_meter': oil.vehicle_meter,
            'maintenance_center': oil.maintenance_center,
            'diesel_filter': oil.diesel_filter,
            'air_filter': oil.air_filter,
            'created_at': oil.created_at
        })
    fuels = Fuel_Filling.objects.filter(Q(vehicle__id=id) & Q(created_at__range=(date_from, date_to))).order_by('created_at')
    fuels_data = []
    for fuel in fuels:
        fuel.created_at = localtime(fuel.created_at, amman_tz)
        fuels_data.append({
            'driver__name': fuel.driver.name,
            'vehicle_meter': fuel.vehicle_meter,
            'litres': fuel.litres,
            'amount': fuel.amount,
            'created_at': fuel.created_at
        })
    oils_count = oils.count()
    fuels_count = fuels.count()
    return JsonResponse({'oils': oils_data, 'fuels': fuels_data, 'oils_count': oils_count, 'fuels_count': fuels_count}, safe=False)

def filter_invoices(request, id):
    date_from = datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d')
    date_to = datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d')
    invoices = Invoices.objects.filter(Q(supplier__id=id) & Q(date__range=(date_from, date_to))).order_by('number')
    amman_tz = pytz.timezone('Asia/Amman')
    invoices_data = []
    for invoice in invoices:
        invoice.created_at = localtime(invoice.created_at, amman_tz)
        invoices_data.append({
            'number': invoice.number,
            'supplier__name': invoice.supplier.name,
            'tamount': invoice.tamount,
            'created_at': invoice.created_at
        })
    treasury = Treasury.objects.filter(Q(invoice__supplier__id=id) & Q(invoice__date__range=(date_from, date_to))).order_by('created_at')
    treasury_data = []
    for tr in treasury:
        tr.created_at = localtime(tr.created_at, amman_tz)
        treasury_data.append({
            'id': tr.id,
            'number': tr.number,
            'inumber': tr.invoice.number,
            'amount': tr.amount,
            'payment': tr.payment,
            'created_at': tr.created_at
        })
    total_invoiced = invoices.aggregate(Sum('tamount'))['tamount__sum'] or 0
    total_treasury = treasury.aggregate(Sum('amount'))['amount__sum'] or 0
    required_amount = total_invoiced - total_treasury
    return JsonResponse({'invoices': invoices_data, 'treasury': treasury_data, 'total_invoiced': total_invoiced, 'total_treasury': total_treasury, 'required_amount': required_amount}, safe=False)

def update_invoice(id):
    invoice = get_object_or_404(Invoices, id=id)
    invoice.tamount = Items.objects.filter(invoice__id=id).aggregate(Sum('tamount'))['tamount__sum'] or 0
    # Extract total_dinar and total_fils
    total_amount = round(invoice.tamount, 2)  # Ensure rounding
    total_dinar = int(total_amount)  # Dinar part
    total_fils = int(round((total_amount - total_dinar) * 100))  # Fils part
    if total_fils > 0:
        invoice.samount = "فقط وقدره " + n2w(total_dinar) + ' دينار و ' + n2w(total_fils) + ' لا غير'
    elif total_dinar > 0:
        invoice.samount = "فقط وقدره " + n2w(total_dinar) + ' دينار لا غير'
    paid = Treasury.objects.filter(invoice__id=id).aggregate(Sum('amount'))['amount__sum'] or 0
    invoice.paid = paid
    invoice.save()

def fetch_attachments(request):
    id = request.GET.get('id')
    diesel_attachments = Fuel_Filling.objects.filter(vehicle__id=id).last()
    oil_attachments = Auto_Oil.objects.filter(vehicle__id=id).last()
    diesel_filter_attachments = Auto_Oil.objects.filter(Q(vehicle__id=id) & Q(diesel_filter=True)).last()
    air_filter_attachments = Auto_Oil.objects.filter(Q(vehicle__id=id) & Q(air_filter=True)).last()
    diesel_attachments_list = []
    oil_attachments_list = []
    diesel_filter_attachments_list = []
    air_filter_attachments_list = []
    if diesel_attachments and diesel_attachments.meter_image:
        diesel_attachments_list.append(
            {'meter_image_url': request.build_absolute_uri(diesel_attachments.meter_image.url)})
    if oil_attachments and oil_attachments.meter_image:
        oil_attachments_list.append({'meter_image_url': request.build_absolute_uri(oil_attachments.meter_image.url)})
    if diesel_filter_attachments and diesel_filter_attachments.meter_image:
        diesel_filter_attachments_list.append(
            {'meter_image_url': request.build_absolute_uri(diesel_filter_attachments.meter_image.url)})
    if air_filter_attachments and air_filter_attachments.meter_image:
        air_filter_attachments_list.append(
            {'meter_image_url': request.build_absolute_uri(air_filter_attachments.meter_image.url)})
    return JsonResponse({
        'diesel_attachments': diesel_attachments_list,
        'oil_attachments': oil_attachments_list,
        'diesel_filter_attachments': diesel_filter_attachments_list,
        'air_filter_attachments': air_filter_attachments_list
    })