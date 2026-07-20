from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, Max, Q, F, Value as V, CharField, DecimalField
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect, FileResponse
from django.urls import reverse
import os
import pytz
from django.utils.timezone import localtime
import openpyxl
from django.conf import settings
import subprocess
import zipfile
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect, FileResponse
from django.contrib import messages
from datetime import datetime, timedelta
from .forms import SearchForm
from .models import Drivers, Treasury, MMission, SMission, PMission
from customers.models import Customers
from recipt_vouchers.models import Vouchers
from invoices.models import Invoices, Items
from invoices.helpers import n2w, update_invoice

# Create your views here.
def index(request):
    new_treasury = Treasury.objects.aggregate(Max('number'))['number__max']
    if new_treasury is None:
        # If no records exist in the table, set the invoice number to 1
        new_treasury = 1
    else:
        # Add 1 to the maximum invoice number
        new_treasury += 1
    drivers = Drivers.objects.all().order_by('-balance')
    form = SearchForm(request.GET)
    result = None
    if "search" in request.GET:
        result = "ok"
        if form.is_valid():
            query = form.cleaned_data['query']
            try:
                query_value = float(query)
                drivers = drivers.filter(Q(number__icontains=query) | Q(balance__gte=query_value))
            except:
                drivers = drivers.filter(Q(name__icontains=query) | Q(username__icontains=query))
    number_of_drivers = drivers.count()
    total_balances = drivers.aggregate(Sum('balance'))['balance__sum'] or 0
    context = {
        'form': form,
        'new_treasury': new_treasury,
        'drivers': drivers,
        'result': result,
        'number_of_drivers': number_of_drivers,
        'total_balances': total_balances,
        'active_tab': 'drivers'  # Set this to the name of the active tab
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('drivers_index')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('drivers_index')
    username = request.user.username
    if 'jobbox' in username:
        return redirect('pmissions_index')
    else:
        return render(request, "drivers/index.html", context)

def get_invoices(request):
    driver_id = request.GET.get('driver_id')
    invoices = Invoices.objects.filter(Q(type="نقدي") & Q(driver__id=driver_id) & Q(tamount__gt=F('driver_paid')))
    invoice_list = list(invoices.values('number'))
    return JsonResponse({'invoices': invoice_list})

def get_invoice_details(request):
    invoice_number = request.GET.get('invoice_number')
    invoice = Invoices.objects.filter(number=invoice_number).first()
    if invoice:
        amount_due = invoice.tamount - invoice.driver_paid
        return JsonResponse({'amount_due': amount_due})
    return JsonResponse({'amount_due': 0})

def driver_balance(request, id):
    driver_name = get_object_or_404(Drivers, id=id).name
    invoices = Invoices.objects.filter(Q(driver__name=driver_name) & Q(type="نقدي"))
    treasury = Treasury.objects.filter(driver__name=driver_name)
    total_invoiced = invoices.aggregate(Sum('tamount'))['tamount__sum'] or 0
    total_treasury = treasury.aggregate(Sum('amount'))['amount__sum'] or 0
    required_amount = total_invoiced - total_treasury
    context = {
        'id': id,
        'driver_name': driver_name,
        'required_amount': required_amount,
        'invoices': invoices,
        'treasury': treasury,
        'active_tab': 'drivers',
        'total_invoiced': total_invoiced,
        'total_treasury': total_treasury
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('drivers_index')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('drivers_index')
    else:
        return render(request, "drivers/balance.html", context)

@login_required
def add_driver(request):
    if request.method == 'POST':
        name = request.POST.get('name').strip()
        username = request.POST.get('username')
        password = request.POST.get('password')
        phone = request.POST.get('phone', '').strip()
        queryset = Drivers.objects.filter(Q(username=username) & Q(password=password))
        if queryset.exists():
            record = queryset.first()  # This will get the first object that matches the query
        else:
            record = None
        # Add the new item to the database
        if record:
            messages.success(request, "Repeated Username and Password...")
        else:
            new_driver = Drivers(name=name, username=username, password=password, phone=phone)
            new_driver.save()
            update_driver_balance(name)
    return redirect('drivers_index')

@login_required
def edit_driver(request, id):
    if request.method == 'POST':
        item = get_object_or_404(Drivers, id=id)
        if item.name != request.POST.get('name').strip():
            vouchers = Vouchers.objects.filter(driver_name=item.name)
            if vouchers:
                for voucher in vouchers:
                    voucher.driver_name = request.POST.get('name').strip()
                    voucher.save()
            item.name = request.POST.get('name').strip()
        item.status = request.POST.get('status')
        item.username = request.POST.get('username')
        item.password = request.POST.get('password')
        item.phone = request.POST.get('phone', '').strip()
        tdriver = Drivers.objects.exclude(id=id)
        if not tdriver:
            item.save()
        else:
            queryset = tdriver.filter(Q(username=request.POST.get('username')) & Q(password=request.POST.get('password')))
            if queryset.exists():
                record = queryset.first()  # This will get the first object that matches the query
            else:
                record = None
            if record:
                messages.success(request, "Repeated Username and Password...")
            else:
                item.save()
    return redirect('drivers_index')

@login_required
def delete_driver(request, id):
    item = get_object_or_404(Drivers, id=id)
    if request.method == 'POST':
        item.delete()
    return redirect('drivers_index')

@login_required
def add_treasury(request):
    if request.method == 'POST':
        driver_id = request.POST.get('driver_id')
        driver = get_object_or_404(Drivers, id=driver_id)
        invoice_number = request.POST.get('invoice_number')
        invoice = get_object_or_404(Invoices, number=invoice_number)
        number = request.POST.get('number')
        amount = request.POST.get('amount')
        payment = request.POST.get('method')
        # Add the new item to the database
        new_treasury = Treasury(driver=driver, invoice=invoice, number=number, amount=amount, payment=payment)
        new_treasury.save()
        update_driver_balance(driver.name)
        update_invoice(invoice.id)
    return redirect('drivers_index')

@login_required
def edit_treasury(request, id):
    if request.method == 'POST':
        item = get_object_or_404(Treasury, id=id)
        item.amount = request.POST.get('edit-amount')
        item.payment = request.POST.get('method')
        item.save()
        driver_name = request.POST.get('driver')
        update_driver_balance(driver_name)
        update_invoice(item.invoice.id)
        related_url = reverse('driver_balance', kwargs={'id': request.POST.get('driver_id')})
        return HttpResponseRedirect(related_url)

@login_required
def delete_treasury(request, id):
    item = get_object_or_404(Treasury, id=id)
    invoice_id = item.invoice.id
    if request.method == 'POST':
        item.delete()
        driver_name = request.POST.get('driver')
        update_driver_balance(driver_name)
        update_invoice(invoice_id)
        related_url = reverse('driver_balance', kwargs={'id': request.POST.get('driver_id')})
        return HttpResponseRedirect(related_url)

@login_required
def export_treasury(request, id, date_from, date_to):
    driver = Drivers.objects.filter(id=id).first()
    driver_name = driver.name
    invoices = Invoices.objects.filter(Q(driver__id=id) & Q(type="نقدي") & (Q(date__range=(date_from, date_to)) | Q(tamount__gt=F('driver_paid')))).annotate(
        account=V('فاتورة', output_field=CharField()),
        number_annotated=F('number'),
        amount_annotated=F('tamount'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'created_at_annotated'
    )

    treasury = Treasury.objects.filter(Q(driver__id=id) & (Q(invoice__date__range=(date_from, date_to)) | Q(invoice__tamount__gt=F('invoice__driver_paid')))).annotate(
        account=V('سداد', output_field=CharField()),
        number_annotated=F('invoice__number'),
        amount_annotated=F('amount'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'created_at_annotated'
    )
    combined_entries = invoices.union(treasury).order_by('created_at_annotated')
    source_excel_path = os.path.join(settings.BASE_DIR, 'excel_templates', 'Driver Account Sheet.xlsx')
    output_excel_path = f"{os.path.splitext(source_excel_path)[0]}_{driver_name}.xlsx"
    # Load the existing workbook (source file)
    wb = openpyxl.load_workbook(source_excel_path)
    ws = wb.active
    cell = ws.cell(row=2, column=1)
    cell.value = "كشف حركة الحساب الخاص بالسائق/ " + driver_name
    white_fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in range(4, 33428):
        ws.row_dimensions[row].hidden = False
    for row in ws.iter_rows(min_row=4, max_row=33428, max_col=4):
        for cell in row:
            cell.value = None
            cell.fill = white_fill
    idx = start_row = 4
    yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    amman_tz = pytz.timezone('Asia/Amman')
    for idx, item in enumerate(combined_entries, start=start_row):
        item['created_at_annotated'] = localtime(item['created_at_annotated'], amman_tz)
        ws.cell(row=idx, column=1, value=item['account'])
        ws.cell(row=idx, column=2, value=item['number_annotated'])
        ws.cell(row=idx, column=3, value=item['amount_annotated'])
        ws.cell(row=idx, column=4, value=item['created_at_annotated'].replace(tzinfo=None))
        if item['account'] == "سداد":
            for col in range(1, 5):  # Assuming you have data in columns 1 to 4
                ws.cell(row=idx, column=col).fill = yellow_fill
    last_data_row = idx + 1
    for value in range(last_data_row, 33428):
        ws.row_dimensions[value].hidden = True
    wb.save(output_excel_path)
    try:
        output_pdf_path = os.path.splitext(output_excel_path)[0] + '.pdf'
        subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf", output_excel_path, '--outdir', os.path.dirname(output_pdf_path)], check=True)
        zip_path = f"{os.path.splitext(output_excel_path)[0]}.zip"
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(output_excel_path, os.path.basename(output_excel_path))
            zipf.write(output_pdf_path, os.path.basename(output_pdf_path))
        if os.path.exists(zip_path):
            with open(zip_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename="Driver_Report_{id}.zip"'
                return response
        else:
            raise Http404("ZIP file not found")
    except Exception as e:
        print("Error:", e)
        raise Http404("An error occurred while generating the files.")

@login_required
def missions_index(request, id):
    drivers = Drivers.objects.all()
    driver = drivers.filter(id=id).first()
    driver_name = driver.name
    missions = MMission.objects.filter(Q(driver__id=id) & Q(receipt__isnull=True))
    missions_count = missions.count()
    customers = Customers.objects.all()
    adjusted_missions = []
    for mission in missions:
        mission.adjusted_count = SMission.objects.filter(Q(mmission__id=mission.id) & Q(receipt="Assigned")).count()
        adjusted_missions.append(mission)
    context = {
        'id': id,
        'drivers': drivers,
        'driver_name': driver_name,
        'missions': adjusted_missions,
        'missions_count': missions_count,
        'customers': customers,
        'active_tab': 'drivers'
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('missions_index')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('missions_index')
    else:
        return render(request, "drivers/missions_index.html", context)

@login_required
def mission_details(request, id):
    mission = get_object_or_404(MMission, id=id)
    items = SMission.objects.filter(Q(mmission__id=id) & Q(receipt="Assigned"))
    driver = mission.driver
    customer = mission.customer
    transportation_date = mission.date
    receiving_location = mission.from_location
    delivery_location = mission.to_location
    estimated_cost = mission.cost
    count_of_vehicles = items.count()
    context = {
               'id': id,
               'items':items,
               'driver':driver,
               'customer':customer,
               'transportation_date':transportation_date,
               'receiving_location':receiving_location,
               'delivery_location':delivery_location,
               'estimated_cost':estimated_cost,
               'count_of_vehicles':count_of_vehicles,
               'active_tab': 'drivers'
               }
    return render(request, "drivers/details.html", context)

@login_required
def add_new_mission(request):
    if request.method == 'POST':
        driver_name = request.POST.get('driver')
        missions_count = MMission.objects.filter(Q(driver__name=driver_name) & Q(date__lte=datetime.today())).count() + 1
        description = driver_name + " " + str(missions_count)
        driver = get_object_or_404(Drivers, name=driver_name)
        customer_name = request.POST.get('name')
        customer = get_object_or_404(Customers, name=customer_name)
        date = request.POST.get('date')
        from_location = request.POST.get('receiving_location')
        to_location = request.POST.get('delivery_location')
        cost = request.POST.get('cost')
        notes = request.POST.get('notes')
        created_by = request.user.username
        new_mission = MMission(description=description, driver=driver, customer=customer, date=date, from_location=from_location, to_location=to_location, cost=cost, scost=get_string(float(cost)), notes=notes, created_by=created_by)
        new_mission.save()
    return redirect('pmissions_index')

@login_required
def add_mission(request, id):
    if request.method == 'POST':
        driver = get_object_or_404(Drivers, id=id)
        customer_name = request.POST.get('name')
        customer_phone = request.POST.get('phone')
        date = request.POST.get('date')
        from_location = request.POST.get('receiving_location')
        to_location = request.POST.get('delivery_location')
        cost = request.POST.get('cost')
        notes = request.POST.get('notes')
        created_by = request.user.username
        queryset = Customers.objects.filter(Q(name=customer_name)|Q(phone=customer_phone))
        if queryset.exists():
            customer = queryset.first()
        else:
            customer = Customers(name=customer_name, phone=customer_phone, created_by=request.user.username)
            customer.save()
        new_mission = MMission(driver=driver, customer=customer, date=date, from_location=from_location, to_location=to_location, cost=cost, scost=get_string(float(cost)), notes=notes, created_by=created_by)
        new_mission.save()
        related_url = reverse('mission_details', kwargs={'id': new_mission.id})
        return HttpResponseRedirect(related_url)

@login_required
def edit_mission(request, id):
    if request.method == 'POST':
        mission = get_object_or_404(MMission, id=id)
        driver = Drivers.objects.filter(name=request.POST.get('driver')).first()
        driver_id = driver.id
        customer_name = request.POST.get('name')
        customer_phone = request.POST.get('phone')
        mission.description = request.POST.get('description')
        mission.driver = driver
        mission.date = request.POST.get('date')
        mission.from_location = request.POST.get('receiving_location')
        mission.to_location = request.POST.get('delivery_location')
        mission.cost = request.POST.get('cost')
        mission.scost = get_string(float(request.POST.get('cost')))
        mission.notes = request.POST.get('notes')
        mission.created_by = request.user.username
        queryset = Customers.objects.filter(Q(name=customer_name) | Q(phone=customer_phone))
        if queryset.exists():
            customer = queryset.first()
        else:
            customer = Customers(name=customer_name, phone=customer_phone)
            customer.save()
        mission.customer = customer
        mission.save()
        related_url = reverse('missions_index', kwargs={'id': driver_id})
        return HttpResponseRedirect(related_url)

@login_required
def delete_mission(request, id):
    if request.method == 'POST':
        mission = get_object_or_404(MMission, id=id)
        driver_id = mission.driver.id
        mission.delete()
        related_url = reverse('missions_index', kwargs={'id': driver_id})
        return HttpResponseRedirect(related_url)

@login_required
def add_mitem(request, id):
    if request.method == 'POST':
        mission = get_object_or_404(MMission, id=id)
        type = request.POST.get('type')
        mark = request.POST.get('mark')
        number = request.POST.get('number').strip()
        color = request.POST.get('color')
        user = request.user.username
        new_item = SMission(mmission=mission, car_type=type, car_mark=mark, car_num=number, car_color=color , created_by=user)
        new_item.save()
        missions_count = SMission.objects.filter(Q(mmission__id=id) & Q(receipt="Assigned")).count()
        mission.count = missions_count
        mission.save()
        related_url = reverse('mission_details', kwargs={'id': mission.id})
        return HttpResponseRedirect(related_url)

@login_required
def edit_mitem(request, id):
    if request.method == 'POST':
        item = get_object_or_404(SMission, id=id)
        item.car_type = request.POST.get('type')
        item.car_mark = request.POST.get('mark')
        item.car_num = request.POST.get('number').strip()
        item.car_color = request.POST.get('color')
        item.created_by = request.user.username
        item.save()
        mission_id = item.mmission.id
        related_url = reverse('mission_details', kwargs={'id': mission_id})
        return HttpResponseRedirect(related_url)

@login_required
def delete_mitem(request, id):
    if request.method == 'POST':
        item = get_object_or_404(SMission, id=id)
        mission = item.mmission
        item.delete()
        missions_count = SMission.objects.filter(Q(mmission__id=item.mmission.id) & Q(receipt="Assigned")).count()
        mission.count = missions_count
        mission.save()
        related_url = reverse('mission_details', kwargs={'id': mission.id})
        return HttpResponseRedirect(related_url)

@login_required
def pmissions_index(request):
    pmissions = PMission.objects.filter(Q(date__gte=datetime.today()) & (Q(mmission__isnull=True) | (Q(mmission__isnull=False) & Q(mmission__count__gt=0) & Q(mmission__accept__isnull=True))))
    mmissions = MMission.objects.filter(date__gte=datetime.today())
    amman_tz = pytz.timezone('Asia/Amman')
    adjusted_pmissions = []
    for pmission in pmissions:
        pmission.created_at = localtime(pmission.created_at, amman_tz)
        pmission.adjusted_created_at = pmission.created_at
        time_diff = pmission.created_at.utcoffset().total_seconds() / 3600
        if pmission.created_at:
            pmission.adjusted_created_at += timedelta(hours=time_diff)
        # Fetch MMissions that match the date and customer of the current PMission
        matching_mmissions = MMission.objects.filter(Q(date=pmission.date) & Q(customer__id=pmission.customer.id) & Q(accept__isnull=True))
        # Attach the options to the PMission object
        pmission.mmission_options = matching_mmissions
        adjusted_pmissions.append(pmission)
    drivers = Drivers.objects.all()
    # Fetch distinct customer IDs from PMission and then get the full Customer objects
    unique_customer_ids = pmissions.values_list('customer', flat=True).distinct()
    customers = Customers.objects.filter(id__in=unique_customer_ids)
    # Fetch distinct from_location and to_location for suggestions
    location_data = {
        customer.id: {
            'from_locations': PMission.objects.filter(customer=customer).values_list('from_location', flat=True).distinct(),
            'to_locations': PMission.objects.filter(customer=customer).values_list('to_location', flat=True).distinct()
        }
        for customer in customers
    }
    context = {
        'mmissions': mmissions,
        'pmissions': adjusted_pmissions,
        'drivers': drivers,
        'customers': customers,
        'location_data': location_data,
        'active_tab': 'pmissions'  # Set this to the name of the active tab
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('pmissions_index')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('pmissions_index')
    return render(request, "drivers/pmissions.html", context)

@login_required
def update_pmission_mmission(request):
    if request.method == 'POST':
        pmission_id = request.POST.get('pmission_id')
        mmission_id = request.POST.get('mmission_id')
        other_notes = request.POST.get('other_notes')
        try:
            pmission = PMission.objects.get(id=pmission_id)
            smission = SMission.objects.filter(Q(car_type=pmission.car_type) & Q(car_num=pmission.car_num.strip()) & Q(mmission__date=pmission.date)).first()
            pmission.another_notes = None
            pmission.save()
            if smission:
                smission.delete()
                mmission = MMission.objects.get(id=pmission.mmission.id)
                missions_count = SMission.objects.filter(Q(mmission__id=mmission.id) & Q(receipt="Assigned")).count()
                mmission.count = missions_count
                mmission.save()
            if mmission_id:
                mmission = MMission.objects.get(id=mmission_id)
                user = request.user.username
                new_item = SMission(mmission=mmission, car_type=pmission.car_type, car_mark=pmission.car_mark, car_num=pmission.car_num.strip(), car_color=pmission.car_color, created_by=user)
                new_item.save()
                pmission.mmission = mmission
                missions_count = SMission.objects.filter(Q(mmission__id=mmission_id) & Q(receipt="Assigned")).count()
                mmission.count = missions_count
                mmission.save()
            else:
                pmission.mmission = None
            pmission.save()
            if other_notes:
                # Save the other notes directly
                pmission.another_notes = other_notes
                pmission.save()
                return JsonResponse({'status': 'success', 'mmission_description': other_notes})
            return JsonResponse({'status': 'success', 'mmission_description': mmission.description if mmission_id else ''})
        except PMission.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'PMission not found'}, status=404)
        except MMission.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'MMission not found'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@login_required
def fetch1(request):
    try:
        name = request.GET.get('name')
        customer = Customers.objects.filter(name=name).first()
        phone = customer.phone
        address = customer.address
    except:
        phone = None
        address = None
    return JsonResponse({'phone': phone, 'address': address})

def get_location_suggestions(request):
    customer_name = request.GET.get('customer_name')
    date = request.GET.get('date')
    if not customer_name or not date:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    # Fetch the customer by name
    try:
        customer = Customers.objects.get(name=customer_name)  # Assumes name is unique
    except Customers.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)
    # Fetch distinct from_location and to_location based on customer and date
    from_locations = PMission.objects.filter(customer=customer, date=date).values_list('from_location', flat=True).distinct()
    to_locations = PMission.objects.filter(customer=customer, date=date).values_list('to_location', flat=True).distinct()
    return JsonResponse({
        'from_locations': list(from_locations),
        'to_locations': list(to_locations),
    })

def filter_invoices(request, id):
    date_from = datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d')
    date_to = datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d')
    invoices = Invoices.objects.filter(Q(driver__id=id) & Q(type="نقدي") & Q(date__range=(date_from, date_to))).order_by('number')
    amman_tz = pytz.timezone('Asia/Amman')
    invoices_data = []
    for invoice in invoices:
        invoice.created_at = localtime(invoice.created_at, amman_tz)
        invoices_data.append({
            'number': invoice.number,
            'require_id__name': invoice.require_id.name,
            'tamount': invoice.tamount,
            'created_at': invoice.created_at
        })
    treasury = Treasury.objects.filter(Q(invoice__driver__id=id) & Q(invoice__type="نقدي") & Q(invoice__date__range=(date_from, date_to))).order_by('created_at')
    treasury_data = []
    for tr in treasury:
        tr.created_at = localtime(tr.created_at, amman_tz)
        treasury_data.append({
            'id': tr.id,
            'number': tr.number,
            'inumber': tr.invoice.number,
            'amount': tr.amount,
            'payment': tr.payment,
            'created_at': tr.created_at
        })
    total_invoiced = invoices.aggregate(Sum('tamount'))['tamount__sum'] or 0
    total_treasury = treasury.aggregate(Sum('amount'))['amount__sum'] or 0
    required_amount = total_invoiced - total_treasury
    return JsonResponse({'invoices': invoices_data, 'treasury': treasury_data, 'total_invoiced': total_invoiced, 'total_treasury': total_treasury, 'required_amount': required_amount}, safe=False)

def fetch_pmissions(request):
    #pmissions = PMission.objects.all()  # Fetch unread missions
    pmissions = PMission.objects.filter(read=False)  # Fetch unread missions
    pmissions_data = []
    for mission in pmissions:
        pmissions_data.append({
            'id': mission.id,
            'customer_name': mission.customer.name,
            'date': mission.date.strftime('%Y-%m-%d') if mission.date else '',
            'from_location': mission.from_location,
            'to_location': mission.to_location,
            'car_type': mission.car_type,
            'car_num': mission.car_num
        })
    return JsonResponse({'pmissions': pmissions_data})

def mark_as_read(request, mission_id):
    if request.method == 'POST':
        try:
            mission = PMission.objects.get(id=mission_id)
            mission.read = True
            mission.save()
            return JsonResponse({'status': 'success', 'message': 'Mission marked as read.'})
        except PMission.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Mission not found.'})

def update_driver_balance(driver_name):
    driver = get_object_or_404(Drivers, name=driver_name)
    total_invoiced = Invoices.objects.filter(Q(driver__name=driver_name) & Q(type="نقدي")).aggregate(Sum('tamount'))['tamount__sum'] or 0
    total_treasury = Treasury.objects.filter(driver__name=driver_name).aggregate(Sum('amount'))['amount__sum'] or 0
    driver.balance = total_invoiced - total_treasury
    driver.save()

def get_string(amount):
    sdinar = int(amount)
    sfils = (amount - int(amount)) * 100
    samount = "فقط وقدره صفر دينار لا غير"
    if sfils > 0:
        samount = "فقط وقدره " + n2w(sdinar) + ' دينار و ' + n2w(sfils) + ' فلس لا غير'
    elif sdinar > 0:
        samount = "فقط وقدره " + n2w(sdinar) + ' دينار لا غير'
    return samount

from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync

def my_view(request):
    # Your view logic
    user = request.user  # Assuming you want to send a message to this user
    message = "A new notification!"

    # Send a message to the WebSocket group for this user
    channel_layer = get_channel_layer()
    group_name = f'user_{user.id}'

    # Use async_to_sync to call async channel layer method from sync view
    async_to_sync(channel_layer.group_send)(
        group_name,
        {
            'type': 'send_notification',  # This corresponds to the method in your consumer
            'message': message,
        }
    )

    return JsonResponse({'status': 'success', 'message': 'Notification sent!'})
