from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, Max, Q, F, Value as V, CharField, DecimalField
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect, FileResponse
from django.urls import reverse
import os
import openpyxl
from django.conf import settings
import subprocess
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect, FileResponse
from django.contrib import messages
import tempfile
from datetime import datetime, date
from drf_yasg.utils import swagger_auto_schema
from django.utils import timezone
from django.core.files.base import ContentFile
import base64
import pytz
import urllib.parse
from openlocationcode import openlocationcode as olc
import requests
from geopy.geocoders import Nominatim
from decimal import Decimal
from openpyxl.drawing.image import Image as OpenpyxlImage
from drivers.views import update_driver_balance
from invoices.helpers import update_invoice
from customers.views import update_customer_balance
from suppliers.views import update_vehicle
from drivers.models import Drivers, Treasury, MMission, SMission
from invoices.models import Invoices, Items
from customers.models import Customers
from recipt_vouchers.models import Vouchers, Attachments
from .models import Used_Vehicle, Auto_Oil, Fuel_Filling
from drivers.serializers import DriverSerializer, TreasurySerializer
from .serializers import VehicleSerializer, LogsSerializer, CustomersSerializer, Vouchers1Serializer, Vouchers2Serializer, AttachmentsSerializer, Vouchers3Serializer, Vouchers31Serializer, Vouchers32Serializer, Vouchers41Serializer, InvoicesSerializer, ItemsSerializer, Items1Serializer, Items2Serializer, InvoicesSerializer2, MMissionSerializer, SMissionSerializer, OilSerializer1, OilSerializer2, FuelSerializer1, FuelSerializer2

# Create your views here.
reference_lat = 31.9539   # Latitude of Amman, Jordan
reference_lng = 35.9106   # Longitude of Amman, Jordan
def reverse_geocode(lat, lon):
    # calling the nominatim tool
    geoLoc = Nominatim(user_agent="GetLoc")
    # passing the coordinates
    locname = geoLoc.reverse(f"{lat}, {lon}", exactly_one=True)
    return locname

def decode_plus_code(plus_code, reference_lat, reference_lng):
    # Convert the short Plus Code to a full Plus Code using a reference location
    full_code = olc.recoverNearest(plus_code, reference_lat, reference_lng)
    # Decode the full Plus Code to get the latitude and longitude
    decoded = olc.decode(full_code)
    return decoded.latitudeCenter, decoded.longitudeCenter

@swagger_auto_schema(
    tags=["Login/logout Page"],
    method='get',
    responses={
        200: DriverSerializer(many=True),
        400: VehicleSerializer(many=True),
    },
    operation_description="Get The data of drivers"
)
@swagger_auto_schema(
    tags=["Login/logout Page"],
    method='post',
    responses={
        200: LogsSerializer(many=True),
    },
    operation_description="Add login data to the database"
)
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def index(request):
    if request.method == "GET":
        drivers = Drivers.objects.all()
        vehicles = Used_Vehicle.objects.filter(Q(status="Available") & Q(driver__isnull=True))
        driver_serializer = DriverSerializer(drivers, many=True)
        vehicle_serializer = VehicleSerializer(vehicles, many=True)
        return Response({
                'drivers': driver_serializer.data,
                'vehicles': vehicle_serializer.data,
            })
    elif request.method == "POST":
        serializer = LogsSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            username = request.data['username']
            password = request.data['password']
            vehicle_number = request.data['vehicle_number']
            driver = Drivers.objects.filter(Q(username=username) & Q(password=password)).first()
            driver_name = driver.name
            vehicle = Used_Vehicle.objects.filter(driver=driver_name).first()
            if vehicle:
                vehicle.driver = None
                vehicle.save()
            vehicle = Used_Vehicle.objects.filter(vehicle_number=vehicle_number).first()
            vehicle.driver = driver_name
            vehicle.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.data, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    tags=["Login/logout Page"],
    method='get',
    operation_description="Logout from the application"
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def logout(request, vehicle_number):
    vehicle = Used_Vehicle.objects.filter(vehicle_number=vehicle_number).first()
    if vehicle:
        vehicle.driver = None
        vehicle.save()
        return Response(status=status.HTTP_200_OK)
    return Response({"error": "Vehicle not found"}, status=status.HTTP_400_BAD_REQUEST)

# Create your views here.
@swagger_auto_schema(
    tags=["Get Customer Data or Add New"],
    method='get',
    responses={
        200: CustomersSerializer(many=True),
    },
    operation_description="Get The data of customers"
)
@swagger_auto_schema(
    tags=["Get Customer Data or Add New"],
    method='post',
    responses={
        200: CustomersSerializer(many=True),
    },
    operation_description="Add new customer"
)
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def get_customer(request, phone):
    if request.method == "GET":
        customer = Customers.objects.filter(phone=phone).first()
        customer_serializer = CustomersSerializer(customer)
        return Response(customer_serializer.data)
    elif request.method == "POST":
        data = request.data.copy()
        data['phone'] = phone
        serializer = CustomersSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    tags=["The First Page"],
    method='get',
    responses={
        200: Vouchers1Serializer(many=True),
    },
    operation_description="Get The data of the customer vouchers"
)
@swagger_auto_schema(
    tags=["The First Page"],
    method='post',
    responses={
        200: Vouchers1Serializer(many=True),
    },
    operation_description="Add new voucher"
)
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def page1_1(request, phone, id):
    if request.method == "GET":
        if id == 0:
            vouchers = Vouchers.objects.filter(Q(recipient_from_name__isnull=True) & Q(car_owner__phone=phone))
            vouchers_serializer = Vouchers1Serializer(vouchers, many=True)
            return Response(vouchers_serializer.data)
        else:
            try:
                smissions = SMission.objects.filter(Q(mmission__id=id) & Q(receipt="Assigned"))
                adjusted_vouchers = []
                for smission in smissions:
                    car_num = smission.car_num
                    voucher = Vouchers.objects.filter(Q(recipient_from_name__isnull=True) & Q(car_owner__phone=phone) & Q(car_num=car_num)).first()
                    if voucher:
                        adjusted_vouchers.append(voucher)
            except SMission.DoesNotExist:
                return Response(status=status.HTTP_404_NOT_FOUND)
            vouchers_serializer = Vouchers1Serializer(adjusted_vouchers, many=True)
            return Response(vouchers_serializer.data)
    elif request.method == "POST":
        data = request.data.copy()
        customer = Customers.objects.filter(phone=phone).first()
        number = Vouchers.objects.aggregate(Max('number'))['number__max']
        if number is None:
            # If no records exist in the table, set the invoice number to 1
            number = 1
        else:
            # Add 1 to the maximum invoice number
            number += 1
        data['number'] = number
        data['car_owner'] = customer.pk
        serializer = Vouchers1Serializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    tags=["The First Page"],
    method='get',
    responses={
        200: Vouchers1Serializer(many=True),
    },
    operation_description="Get The data of the selected voucher"
)
@swagger_auto_schema(
    tags=["The First Page"],
    method='put',
    responses={
        200: Vouchers1Serializer(many=True),
    },
    operation_description="Edit the data of the selected voucher"
)
@swagger_auto_schema(
    tags=["The First Page"],
    method='delete',
    operation_description="Delete the data of the selected voucher"
)
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def page1_2(request, phone, number):
    customer = Customers.objects.filter(phone=phone).first()
    vouchers = Vouchers.objects.filter(number=number).first()
    if request.method == "GET":
        vouchers_serializer = Vouchers1Serializer(vouchers)
        return Response(vouchers_serializer.data)
    elif request.method == "PUT":
        data = request.data.copy()
        data['number'] = number
        data['car_owner'] = customer.pk
        serializer = Vouchers1Serializer(vouchers, data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.data, status=status.HTTP_400_BAD_REQUEST)
    elif request.method == "DELETE":
        vouchers.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

@swagger_auto_schema(
    tags=["The Second Page"],
    method='get',
    responses={
        200: Vouchers2Serializer(many=True),
    },
    operation_description="Get The data of the last voucher"
)
@swagger_auto_schema(
    tags=["The Second Page"],
    method='put',
    responses={
        200: Vouchers2Serializer(many=True),
    },
    operation_description="Edit the data of the last voucher"
)
@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def page2_1(request, number):
    vouchers = Vouchers.objects.filter(number=number).first()
    if not vouchers:
        return Response({'error': 'Voucher not found'}, status=status.HTTP_404_NOT_FOUND)
    if request.method == "GET":
        vouchers_serializer = Vouchers2Serializer(vouchers)
        return Response(vouchers_serializer.data)
    elif request.method == "PUT":
        data = request.data.copy()
        vouchers.receipt_from_time = datetime.now()
        vouchers.save()
        try:
            latitude, longitude = decode_plus_code(data['receipt_from_location'], reference_lat, reference_lng)
            data['receipt_from_location'] = reverse_geocode(latitude, longitude).address
        except:
            pass
        if data['notes']:
            data['notes'] = "ملاحظات الاستلام: " + data['notes']
        serializer = Vouchers2Serializer(vouchers, data=data)  # Use partial=True for partial updates
        if serializer.is_valid():
            serializer.save()
            item = SMission.objects.filter(Q(car_num=vouchers.car_num) & Q(receipt="Assigned")).first()
            if item:
                item.receipt = "Received"
                item.save()
                item.mmission.accept = "Done"
                item.mmission.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    tags=["The Second Page"],
    method='get',
    responses={
        200: AttachmentsSerializer(many=True),
    },
    operation_description="Get The attachments of the Page2"
)
@swagger_auto_schema(
    tags=["The Second Page"],
    method='post',
    responses={
        200: AttachmentsSerializer(many=True),
    },
    operation_description="Add new attachments to Page2"
)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def page2_2(request, number):
    vouchers = Vouchers.objects.filter(number=number).first()
    if not vouchers:
        return Response({'error': 'Voucher not found'}, status=status.HTTP_404_NOT_FOUND)
    if request.method == "GET":
        attachments = Attachments.objects.filter(voucher=vouchers)
        attachments_serializer = AttachmentsSerializer(attachments, many=True)
        return Response(attachments_serializer.data)
    elif request.method == "POST":
        images = request.FILES.getlist('strike_chart')  # Fetch the list of images
        if not images:
            return Response({'error': 'No images provided'}, status=status.HTTP_400_BAD_REQUEST)
        saved_attachments = []
        for image in images:
            attachment_data = {
                'voucher': vouchers.pk,
                'strike_chart': image,  # Assign each image to strike_chart field
                'page': "Received"
            }
            serializer = AttachmentsSerializer(data=attachment_data)
            if serializer.is_valid():
                saved_attachments.append(serializer.save())
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        attachments_serializer = AttachmentsSerializer(saved_attachments, many=True)
        return Response(attachments_serializer.data, status=status.HTTP_201_CREATED)

@swagger_auto_schema(
    tags=["The Third Page"],
    method='get',
    responses={
        200: Vouchers3Serializer(many=True),
    },
    operation_description="Get The data of the undelivered vehicles vouchers"
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def page3(request, driver_name):
    decoded_driver_name = urllib.parse.unquote(driver_name)
    vouchers = Vouchers.objects.filter(Q(driver_name=decoded_driver_name) & Q(recipient_to_name__isnull=True) & Q(recipient_to_name__isnull=True) & Q(recipient_from_name__isnull=False)).order_by('receipt_from_time')
    if request.method == "GET":
        vouchers_serializer = Vouchers3Serializer(vouchers, many=True)
        return Response(vouchers_serializer.data)

@swagger_auto_schema(
    tags=["The Third Page"],
    method='get',
    responses={
        200: Vouchers31Serializer(many=True),
    },
    operation_description="Get The data of the selected voucher"
)
@swagger_auto_schema(
    tags=["The Third Page"],
    method='put',
    responses={
        200: Vouchers32Serializer(many=True),
    },
    operation_description="Edit the data of the selected voucher"
)
@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def page3_1(request, number):
    try:
        vouchers = Vouchers.objects.filter(number=number).first()
        if not vouchers:
            return Response({'error': 'Voucher not found'}, status=status.HTTP_404_NOT_FOUND)
        if request.method == "GET":
            vouchers_serializer = Vouchers31Serializer(vouchers)
            return Response(vouchers_serializer.data)
        elif request.method == "PUT":
            data = request.data.copy()
            vouchers.receipt_to_time = datetime.now()
            vouchers.save()
            try:
                latitude, longitude = decode_plus_code(data['receipt_to_location'], reference_lat, reference_lng)
                data['receipt_to_location'] = reverse_geocode(latitude, longitude).address
            except:
                pass
            req = data['notes']
            if vouchers.notes:
                data['notes'] = vouchers.notes
                if req:
                    if req != vouchers.notes[18:]:
                        data['notes'] += " - ملاحظات التسليم: " + req
            elif req:
                data['notes'] = "ملاحظات التسليم: " + req
            serializer = Vouchers32Serializer(vouchers, data=data)
            if serializer.is_valid():
                serializer.save()
                item = SMission.objects.filter(Q(car_num=vouchers.car_num) & Q(receipt="Received")).first()
                if item:
                    item.receipt = "Delivered"
                    item.save()
                    mission = get_object_or_404(MMission, id=item.mmission.id)
                    items_count = SMission.objects.filter(Q(mmission__id=item.mmission.id) & Q(receipt="Received")).count()
                    mission.count = items_count
                    if items_count == 0:
                        mission.receipt = "Delivered"
                    mission.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except:
        return Response({'error': 'Server error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@swagger_auto_schema(
    tags=["The Third Page"],
    method='get',
    responses={
        200: AttachmentsSerializer(many=True),
    },
    operation_description="Get The attachments of the Page3"
)
@swagger_auto_schema(
    tags=["The Third Page"],
    method='post',
    responses={
        200: AttachmentsSerializer(many=True),
    },
    operation_description="Add new attachments to Page3"
)
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def page3_2(request, number):
    vouchers = Vouchers.objects.filter(number=number).first()
    attachments = Attachments.objects.filter(voucher=vouchers)
    if request.method == "GET":
        attachments_serializer = AttachmentsSerializer(attachments, many=True)  # Use AttachmentsSerializer here
        return Response(attachments_serializer.data)
    elif request.method == "POST":
        images = request.FILES.getlist('strike_chart')  # Fetch the list of images
        if not images:
            return Response({'error': 'No images provided'}, status=status.HTTP_400_BAD_REQUEST)
        saved_attachments = []
        for image in images:
            attachment_data = {
                'voucher': vouchers.pk,
                'strike_chart': image,  # Assign each image to strike_chart field
                'page': "Delivered"
            }
            serializer = AttachmentsSerializer(data=attachment_data)
            if serializer.is_valid():
                saved_attachments.append(serializer.save())
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        attachments_serializer = AttachmentsSerializer(saved_attachments, many=True)
        return Response(attachments_serializer.data, status=status.HTTP_201_CREATED)

@swagger_auto_schema(
    tags=["The Third Page"],
    method='get',
    operation_description="Get The number of residual vehicles"
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def page3_3(request, phone, driver_name):
    decoded_driver_name = urllib.parse.unquote(driver_name)
    customer = Customers.objects.filter(phone=phone).first()
    if not customer:
        return Response({'error': 'Customer not found'}, status=status.HTTP_404_NOT_FOUND)
    residual = Vouchers.objects.filter(Q(car_owner__phone=phone) & Q(driver_name=decoded_driver_name) & Q(recipient_to_name__isnull=True) & Q(recipient_from_name__isnull=False)).count()
    if residual > 0:
        return Response({'Alert': 'Next Button', 'Count': residual}, status=status.HTTP_200_OK)
    else:
        return Response({'Alert': 'Invoice Button', 'Count': residual}, status=status.HTTP_200_OK)

@swagger_auto_schema(
    tags=["The Fourth Page"],
    method='get',
    responses={
        200: Vouchers41Serializer(many=True),
    },
    operation_description="Get The Vehicles that are without invoices"
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def page4(request, phone):
    vouchers = Vouchers.objects.filter(Q(invoice__isnull=True) & Q(car_owner__phone=phone) & Q(recipient_to_name__isnull=False))
    if request.method == "GET":
        vouchers_serializer = Vouchers41Serializer(vouchers, many=True)
        return Response(vouchers_serializer.data)

@swagger_auto_schema(
    tags=["The Fourth Page"],
    method='get',
    responses={
        200: InvoicesSerializer(many=True),
    },
    operation_description="Get The last invoice for the customer by his phone"
)
@swagger_auto_schema(
    tags=["The Fourth Page"],
    method='put',
    responses={
        200: InvoicesSerializer2(many=True),
    },
    operation_description="Edit the data of the last invoice for the customer by his phone"
)
@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def get_invoices(request, phone, driver_name, vehicle_number):
    decoded_driver_name = urllib.parse.unquote(driver_name)
    customer = Customers.objects.filter(phone=phone).first()
    if not customer:
        return Response({'error': 'Customer not found'}, status=status.HTTP_404_NOT_FOUND)
    vehicle = Used_Vehicle.objects.filter(vehicle_number=vehicle_number).first()
    invoice = Invoices.objects.filter(Q(require_id__id=customer.id) & Q(driver__name=decoded_driver_name) & Q(vehicle__id=vehicle.id)).last()
    if not invoice:
        return Response({'error': 'No invoices available'}, status=status.HTTP_404_NOT_FOUND)
    if invoice.exported == "Done":
        return Response({'error': 'This invoice is shared before'}, status=status.HTTP_404_NOT_FOUND)
    items = Items.objects.filter(number__id=invoice.id)
    if request.method == "GET":
        invoice_serializer = InvoicesSerializer(invoice)
        items_serializer = ItemsSerializer(items, many=True)
        return Response({
            'invoice': invoice_serializer.data,
            'items': items_serializer.data,
        })
    elif request.method == "PUT":
        serializer = InvoicesSerializer2(invoice, data=request.data)
        if request.data['type'] == 'كليك' or request.data['type'] == 'نقدي' or (request.data['type'] == 'ذمم' and customer.payment == 'ذمم'):
            if serializer.is_valid():
                serializer.save()
                update_driver_balance(decoded_driver_name)
                update_customer_balance(customer.id)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response({'error': 'Not Available Payment Method for the Customer'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.data, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    tags=["The Fourth Page"],
    method='get',
    responses={
        200: ItemsSerializer(many=True),
    },
    operation_description="Get The items of the last invoice for the customer by his phone"
)
@swagger_auto_schema(
    tags=["The Fourth Page"],
    method='post',
    responses={
        200: Items2Serializer(many=True),
    },
    operation_description="Add new item to the last invoice for the customer by his phone"
)
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def get_items(request, phone, driver_name, vehicle_number):
    decoded_driver_name = urllib.parse.unquote(driver_name)
    customer = Customers.objects.filter(phone=phone).first()
    invoice = Invoices.objects.filter(Q(require_id__id=customer.id) & Q(driver__name=decoded_driver_name)).last()
    try:
        items = Items.objects.filter(number__id=invoice.id)
    except Items.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    if request.method == "GET":
        serializer = ItemsSerializer(items, many=True)
        return Response(serializer.data)
    elif request.method == "POST":
        data = request.data.copy()
        data['number'] = invoice.pk
        serializer = Items2Serializer(data=data)
        if serializer.is_valid():
            serializer.save()
            item = Items.objects.filter(number=invoice.pk).last()
            # Calculate adinar and afils
            total_price = round((item.udinar + (item.ufils / 100)) * item.qty, 2)
            item.adinar = int(total_price)
            item.afils = int(round((total_price - item.adinar) * 100))
            item.save()
            update_invoice(invoice.pk)
            update_driver_balance(decoded_driver_name)
            update_vehicle(vehicle_number)
            update_customer_balance(customer.id)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.data, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    tags=["The Fourth Page"],
    method='get',
    responses={
        200: ItemsSerializer(many=True),
    },
    operation_description="Get an item by its pk"
)
@swagger_auto_schema(
    tags=["The Fourth Page"],
    method='put',
    responses={
        200: Items1Serializer(many=True),
    },
    operation_description="Edit the data of an item by its pk"
)
@swagger_auto_schema(
    tags=["The Fourth Page"],
    method='delete',
    operation_description="Delete the data of an item by its pk"
)
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def put_item(request, phone, driver_name, vehicle_number, pk):
    decoded_driver_name = urllib.parse.unquote(driver_name)
    customer = Customers.objects.filter(phone=phone).first()
    invoice = Invoices.objects.filter(Q(require_id__id=customer.id) & Q(driver__name=decoded_driver_name)).last()
    try:
        item = Items.objects.filter(id=pk).first()
    except:
        item = None
    if request.method == "GET":
        serializer = ItemsSerializer(item)
        return Response(serializer.data)
    elif request.method == "PUT":
        # Parsing and calculating the total value in dinars and fils
        unit_price = float(request.data['udinar']) + (float(request.data['ufils']) / 100)
        total_price = round(unit_price * float(request.data['qty']), 2)
        # Calculating adinar (integer part) and afils (integer part for fils)
        item.adinar = int(total_price)
        item.afils = int(round((total_price - item.adinar) * 100))
        item.save()
        serializer = Items1Serializer(item, data=request.data)
        if serializer.is_valid():
            serializer.save()
            update_invoice(invoice.pk)
            update_driver_balance(decoded_driver_name)
            update_vehicle(vehicle_number)
            update_customer_balance(customer.id)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.data, status=status.HTTP_400_BAD_REQUEST)
    elif request.method == "DELETE":
        item.delete()
        update_invoice(invoice.pk)
        update_driver_balance(decoded_driver_name)
        update_vehicle(vehicle_number)
        update_customer_balance(customer.id)
        return Response(status=status.HTTP_204_NO_CONTENT)

@swagger_auto_schema(
    tags=["The Fourth Page"],
    method='get',
    operation_description="Export The invoices as pdf"
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_invoice(request, phone, driver_name, vehicle_number):
    decoded_driver_name = urllib.parse.unquote(driver_name)
    customer = Customers.objects.filter(phone=phone).first()
    if not customer:
        return Response({'error': 'Customer not found'}, status=status.HTTP_404_NOT_FOUND)
    pre_vouchers = Vouchers.objects.filter(Q(recipient_to_name__isnull=False) & Q(car_owner__phone=phone) & Q(driver_name=decoded_driver_name))
    if not pre_vouchers:
        return Response({'error': 'Vouchers not found'}, status=status.HTTP_404_NOT_FOUND)
    vouchers = pre_vouchers.filter(invoice__isnull=True)
    vehicle = Used_Vehicle.objects.filter(vehicle_number=vehicle_number).first()
    driver = Drivers.objects.filter(name=decoded_driver_name).first()
    if vouchers:
        number = Invoices.objects.aggregate(Max('number'))['number__max']
        if number is None:
            # If no records exist in the table, set the invoice number to 1
            number = 1
        else:
            # Add 1 to the maximum invoice number
            number += 1
        cars = vouchers.values_list('car_num', flat=True)
        # Add the new item to the database
        new_invoice = Invoices(driver=driver, number=number, type=customer.payment, date=date.today(), require_id=customer, car=', '.join(cars), notes="", created_at=datetime.now(), vehicle=vehicle)
        new_invoice.save()
        qty = 0
        adinar = 0
        afils = 0
        amount = 0
        for voucher in vouchers:
            qty += 1
            voucher.invoice = new_invoice
            voucher.save()
            from_location = voucher.receipt_from_location
            to_location = voucher.receipt_to_location
            item = SMission.objects.filter(Q(car_num=voucher.car_num) & Q(receipt="Delivered") & Q(mmission__driver__name=decoded_driver_name)).first()
            if item:
                item.receipt = "Invoiced"
                item.save()
                mission = get_object_or_404(MMission, id=item.mmission.id)
                from_location = mission.from_location
                to_location = mission.to_location
                adinar = int(mission.cost)
                afils = int(round((mission.cost - adinar) * 100))
                new_invoice.dinar = adinar
                new_invoice.fils = afils
                amount = mission.cost
                new_invoice.tamount = amount
                new_invoice.samount = mission.scost
                new_invoice.save()
            description = f"نقل مركبات للعميل من {from_location} إلى {to_location}"
        udinar = int(amount / qty)
        ufils = int(round(((amount / qty) - udinar) * 100))
        new_item = Items(number=new_invoice, description=description, qty=qty, udinar=udinar, ufils=ufils, adinar=adinar, afils=afils)
        new_item.save()
        update_driver_balance(decoded_driver_name)
        update_vehicle(vehicle_number)
        update_customer_balance(customer.id)
    invoice = Invoices.objects.filter(Q(require_id__id=customer.id) & Q(driver__name=decoded_driver_name) & Q(vehicle__id=vehicle.id)).last()
    if not invoice:
        return Response({'error': 'Invoice not found'}, status=status.HTTP_404_NOT_FOUND)
    items = Items.objects.filter(number__id=invoice.id)
    customer_name = customer.name
    sub_customer = ""
    if customer.notes:
        if "شرك" in customer.notes and "مركز" in customer.notes and "تجار" in customer.notes and "مركبات" in customer.notes:
            sub_customer = customer_name
            customer_name = customer.notes
    voucher = Vouchers.objects.filter(invoice__id=invoice.id).first()
    recipient_to_name = voucher.recipient_to_name
    image_path = voucher.receipt_to_signature.path
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Invoice Number ' + customer_name + '.xlsx"'
    source_excel_path = os.path.join(settings.BASE_DIR, 'excel_templates', 'invoice.xlsx')
    output_excel_path = f"{os.path.splitext(source_excel_path)[0]} Customer: {customer_name}.xlsx"
    # Load the existing workbook (source file)
    wb = openpyxl.load_workbook(source_excel_path)
    ws = wb.active
    img = OpenpyxlImage(image_path)
    ws.add_image(img, 'B33')
    # Fetch data from the model
    ws['A6'] = invoice.number
    ws['E7'] = invoice.type
    ws['G8'] = invoice.date
    ws['B10'] = customer_name
    ws['A12'] = invoice.car
    ws['B30'] = invoice.samount
    ws['H31'] = invoice.dinar
    ws['I31'] = invoice.fils
    ws['E33'] = recipient_to_name
    ws['E34'] = sub_customer
    ws['E35'] = invoice.driver.name
    ws['E37'] = invoice.notes
    for idx, item in enumerate(items, start=18):
        ws.cell(row=idx, column=1, value=item.udinar)
        ws.cell(row=idx, column=2, value=item.ufils)
        ws.cell(row=idx, column=3, value=item.qty)
        ws.cell(row=idx, column=4, value=item.description)
        ws.cell(row=idx, column=8, value=item.adinar)
        ws.cell(row=idx, column=9, value=item.afils)
    wb.save(output_excel_path)
    try:
        output_pdf_path = os.path.splitext(output_excel_path)[0] + '.pdf'
        subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf", output_excel_path, '--outdir', os.path.dirname(output_pdf_path)], check=True)
        # Provide the PDF as a download
        if os.path.exists(output_pdf_path):
            with open(output_pdf_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{os.path.basename(output_pdf_path)}"'
                return response
        else:
            raise Http404("PDF file not found")
    except Exception as e:
        print("Error:", e)

@swagger_auto_schema(
    tags=["The Fourth Page"],
    method='get',
    operation_description="Share the invoice to the customer"
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def share_invoice(request, phone, driver_name, vehicle_number):
    decoded_driver_name = urllib.parse.unquote(driver_name)
    customer = Customers.objects.filter(phone=phone).first()
    vehicle = Used_Vehicle.objects.filter(vehicle_number=vehicle_number).first()
    invoice = Invoices.objects.filter(Q(require_id__id=customer.id) & Q(driver__name=decoded_driver_name) & Q(vehicle__id=vehicle.id)).last()
    invoice.exported = "Done"
    invoice.save()
    return Response({'Success': 'Shared...'}, status=status.HTTP_200_OK)

@swagger_auto_schema(
    tags=["Todo Page"],
    method='get',
    responses={
        200: MMissionSerializer(many=True),
    },
    operation_description="Get The Mission for the Driver"
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_missions(request, driver_name):
    decoded_driver_name = urllib.parse.unquote(driver_name)
    try:
        missions = MMission.objects.filter(Q(driver__name=decoded_driver_name) & Q(count__gt=0))
    except MMission.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    if request.method == "GET":
        serializer = MMissionSerializer(missions, many=True)
        return Response(serializer.data)

@swagger_auto_schema(
    tags=["Todo Page"],
    method='get',
    responses={
        200: SMissionSerializer(many=True),
    },
    operation_description="Get The Items of Required Mission"
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_sub_missions(request, id):
    try:
        smissions = SMission.objects.filter(Q(mmission__id=id) & Q(receipt="Assigned"))
    except SMission.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    if request.method == "GET":
        serializer = SMissionSerializer(smissions, many=True)
        return Response(serializer.data)

@swagger_auto_schema(
    tags=["The Oil Page"],
    method='post',
    responses={
        200: OilSerializer1(many=True),
    },
    operation_description="add new record about the vehicle oil"
)
@swagger_auto_schema(
    tags=["The Oil Page"],
    method='get',
    responses={
        200: OilSerializer2(many=True),
    },
    operation_description="Get the last status of vehicle oil"
)
@swagger_auto_schema(
    tags=["The Oil Page"],
    method='put',
    responses={
        200: OilSerializer2(many=True),
    },
    operation_description="Edit the last status of vehicle oil"
)
@api_view(['POST', 'GET', 'PUT'])
@permission_classes([IsAuthenticated])
def get_oil(request, driver_name, vehicle_number):
    decoded_driver_name = urllib.parse.unquote(driver_name)
    driver = Drivers.objects.filter(name=decoded_driver_name).first()
    vehicle = Used_Vehicle.objects.filter(vehicle_number=vehicle_number).first()
    oil_status = Auto_Oil.objects.filter(vehicle__id=vehicle.id).last()
    if request.method == "GET":
        serializer = OilSerializer2(oil_status)
        return Response(serializer.data)
    elif request.method == "PUT":
        exoil = Auto_Oil.objects.filter(vehicle__id=vehicle.id).exclude(id=oil_status.id).last()
        if Decimal(request.data['vehicle_meter']) < vehicle.fuel_meter:
            return Response({'error': 'The meter difference must be greater than the last fuel meter'}, status=status.HTTP_400_BAD_REQUEST)
        if exoil and Decimal(request.data['vehicle_meter']) < exoil.vehicle_meter + 4000:
            return Response({
                'error': 'The meter difference must be more than or equal to 4000 kms',
                'previous_oil_meter': exoil.vehicle_meter  # Include exoil.vehicle_meter in the response
            }, status=status.HTTP_400_BAD_REQUEST)
        serializer = OilSerializer2(oil_status, data=request.data)
        if serializer.is_valid():
            serializer.save()
            vehicle.oil_meter = request.data['vehicle_meter']
            vehicle.current_oil_diff = 0.0
            if request.data.get('air_filter') is True:
                vehicle.air_filter_meter = request.data['vehicle_meter']
                vehicle.current_air_filter_diff = 0.0
            if request.data.get('diesel_filter') is True:
                vehicle.fuel_filter_meter = request.data['vehicle_meter']
                vehicle.current_fuel_filter_diff = 0.0
            vehicle.save()
            oil_status.driver = driver
            oil_status.created_at = datetime.now()
            oil_status.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    elif request.method == "POST":
        data = request.data.copy()
        if Decimal(data['vehicle_meter']) < vehicle.fuel_meter:
            return Response({'error': 'The meter difference must be greater than the last fuel meter'}, status=status.HTTP_400_BAD_REQUEST)
        if Decimal(data['vehicle_meter']) < vehicle.oil_meter + 4000:
            return Response({
                'error': 'The meter difference must be more than or equal to 4000 kms',
                'previous_oil_meter': vehicle.oil_meter  # Include exoil.vehicle_meter in the response
            }, status=status.HTTP_400_BAD_REQUEST)
        # Check if both air filter and diesel filter need to be changed
        if not data.get('air_filter') and not data.get('diesel_filter') and vehicle.current_air_filter_diff >= vehicle.default_air_filter_diff and vehicle.current_fuel_filter_diff >= vehicle.default_fuel_filter_diff:
            return Response({'error': 'You must change both the air filter and the diesel filter'}, status=status.HTTP_400_BAD_REQUEST)
        # Check if only the air filter needs to be changed
        if not data.get('air_filter') and vehicle.current_air_filter_diff >= vehicle.default_air_filter_diff:
            return Response({'error': 'You must change the air filter'}, status=status.HTTP_400_BAD_REQUEST)
        # Check if only the diesel filter needs to be changed
        if not data.get('diesel_filter') and vehicle.current_fuel_filter_diff >= vehicle.default_fuel_filter_diff:
            return Response({'error': 'You must change the diesel filter'}, status=status.HTTP_400_BAD_REQUEST)
        data['driver'] = driver.pk
        data['vehicle'] = vehicle.pk
        serializer = OilSerializer1(data=data)
        if serializer.is_valid():
            serializer.save()
            vehicle.oil_meter = data['vehicle_meter']
            vehicle.current_oil_diff = 0.0
            if data.get('air_filter') is True:
                vehicle.air_filter_meter = data['vehicle_meter']
                vehicle.current_air_filter_diff = 0.0
            if data.get('diesel_filter') is True:
                vehicle.fuel_filter_meter = data['vehicle_meter']
                vehicle.current_fuel_filter_diff = 0.0
            vehicle.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    tags=["The Fuel Page"],
    method='post',
    responses={
        200: FuelSerializer1(many=True),
    },
    operation_description="add new record about the vehicle fuel"
)
@swagger_auto_schema(
    tags=["The Fuel Page"],
    method='get',
    responses={
        200: FuelSerializer2(many=True),
    },
    operation_description="Get the last status of vehicle fuel"
)
@swagger_auto_schema(
    tags=["The Fuel Page"],
    method='put',
    responses={
        200: FuelSerializer2(many=True),
    },
    operation_description="Edit the last status of vehicle fuel"
)
@api_view(['POST', 'GET', 'PUT'])
@permission_classes([IsAuthenticated])
def get_fuel(request, driver_name, vehicle_number):
    decoded_driver_name = urllib.parse.unquote(driver_name)
    driver = Drivers.objects.filter(name=decoded_driver_name).first()
    vehicle = Used_Vehicle.objects.filter(vehicle_number=vehicle_number).first()
    fuel_status = Fuel_Filling.objects.filter(vehicle__id=vehicle.id).last()
    if request.method == "GET":
        serializer = FuelSerializer2(fuel_status)
        return Response(serializer.data)
    elif request.method == "PUT":
        exfuel = Fuel_Filling.objects.filter(vehicle__id=vehicle.id).exclude(id=fuel_status.id).last()
        if exfuel and Decimal(request.data['vehicle_meter']) < exfuel.vehicle_meter:
            return Response({'error': 'The meter difference must be greater than the last fuel meter'}, status=status.HTTP_400_BAD_REQUEST)
        if Decimal(request.data['vehicle_meter']) < vehicle.oil_meter:
            return Response({'error': 'The meter difference must be greater than the last oil meter'}, status=status.HTTP_400_BAD_REQUEST)
        serializer = FuelSerializer2(fuel_status, data=request.data)
        if serializer.is_valid():
            serializer.save()
            vehicle.fuel_meter = request.data['vehicle_meter']
            vehicle.current_oil_diff = Decimal(request.data['vehicle_meter']) - vehicle.oil_meter
            vehicle.current_air_filter_diff = Decimal(request.data['vehicle_meter']) - vehicle.air_filter_meter
            vehicle.current_fuel_filter_diff = Decimal(request.data['vehicle_meter']) - vehicle.fuel_filter_meter
            vehicle.save()
            fuel_status.driver = driver
            fuel_status.created_at = datetime.now()
            fuel_status.save()
            response_data = serializer.data
            if vehicle.current_oil_diff >= vehicle.default_oil_diff:
                response_data['alert'] = 'Warning: Current oil difference exceeds the default limit!'
            return Response(response_data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    elif request.method == "POST":
        data = request.data.copy()
        if Decimal(data['vehicle_meter']) < vehicle.fuel_meter:
            return Response({'error': 'The meter difference must be greater than the last fuel meter'}, status=status.HTTP_400_BAD_REQUEST)
        if Decimal(data['vehicle_meter']) < vehicle.oil_meter:
            return Response({'error': 'The meter difference must be greater than the last oil meter'}, status=status.HTTP_400_BAD_REQUEST)
        data['driver'] = driver.pk
        data['vehicle'] = vehicle.pk
        serializer = FuelSerializer1(data=data)
        if serializer.is_valid():
            serializer.save()
            vehicle.fuel_meter = data['vehicle_meter']
            vehicle.current_oil_diff = Decimal(data['vehicle_meter']) - vehicle.oil_meter
            vehicle.current_air_filter_diff = Decimal(data['vehicle_meter']) - vehicle.air_filter_meter
            vehicle.current_fuel_filter_diff = Decimal(data['vehicle_meter']) - vehicle.fuel_filter_meter
            vehicle.save()
            response_data = serializer.data
            if vehicle.current_oil_diff >= vehicle.default_oil_diff:
                response_data['alert'] = 'Warning: Current oil difference exceeds the default limit!'
            return Response(response_data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    tags=["The Oil Page"],
    method='get',
    operation_description="Get the necessary alerts with oil page"
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_alerts(request, vehicle_number):
    vehicle = Used_Vehicle.objects.filter(vehicle_number=vehicle_number).first()
    if not vehicle:
        return Response({'error': 'Vehicle not found'}, status=status.HTTP_404_NOT_FOUND)
    if request.method == "GET":
        # Check if both air filter and diesel filter need to be changed
        if vehicle.current_air_filter_diff >= vehicle.default_air_filter_diff and vehicle.current_fuel_filter_diff >= vehicle.default_fuel_filter_diff:
            return Response({'Alert': 'You must change both the air filter and the diesel filter'}, status=status.HTTP_200_OK)
        # Check if only the air filter needs to be changed
        if vehicle.current_air_filter_diff >= vehicle.default_air_filter_diff:
            return Response({'Alert': 'You must change the air filter'}, status=status.HTTP_200_OK)
        # Check if only the diesel filter needs to be changed
        if vehicle.current_fuel_filter_diff >= vehicle.default_fuel_filter_diff:
            return Response({'Alert': 'You must change the diesel filter'}, status=status.HTTP_200_OK)
        # If no filters need changing, return success
        return Response({'Success': 'There are no alerts'}, status=status.HTTP_200_OK)


