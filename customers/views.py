from django.shortcuts import render, redirect, get_object_or_404
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, Max, Q, F, Value as V, CharField, DecimalField, OuterRef, Subquery
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect, FileResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
import os
import openpyxl
from django.conf import settings
import subprocess
import zipfile
from django.core.files.storage import FileSystemStorage
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
import pytz
import pandas as pd
from django.utils.timezone import localtime
from datetime import datetime, timedelta
from django.contrib.auth.models import User
from .forms import SearchForm, ExcelUploadForm
from recipt_vouchers.models import Vouchers
from .models import Customers, Papers, Treasury
from invoices.models import Invoices
from invoices.helpers import update_invoice

def index(request):
    new_treasury = Treasury.objects.aggregate(Max('number'))['number__max']
    if new_treasury is None:
        # If no records exist in the table, set the invoice number to 1
        new_treasury = 1
    else:
        # Add 1 to the maximum invoice number
        new_treasury += 1
    customers = Customers.objects.filter(balance__gt=0).order_by('-balance')
    form = SearchForm(request.GET)
    result = None
    options = "option0"
    if "search" in request.GET:
        result = "ok"
        if form.is_valid():
            if request.GET['inlineRadioOptions'] == "option1":
                customers = Customers.objects.filter(payment="ذمم").order_by('-balance')
                options = "option1"
            elif request.GET['inlineRadioOptions'] == "option2":
                customers = Customers.objects.filter(payment="نقدي").order_by('-balance')
                options = "option2"
            query = form.cleaned_data['query']
            customers = customers.filter(Q(name__icontains=query) | Q(phone__icontains=query) | Q(mail__icontains=query) | Q(notes__icontains=query) | Q(username__icontains=query))
    number_of_customers = customers.count()
    total_balances = customers.aggregate(Sum('balance'))['balance__sum'] or 0
    context = {
        'active_tab': 'customers',
        'new_treasury': new_treasury,
        'result': result,
        'options': options,
        'form': form,
        'customers': customers,
        'number_of_customers': number_of_customers,
        'total_balances': total_balances
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('customers_index')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('customers_index')
    else:
        username = request.user.username
        if 'jobbox' in username:
            return redirect('pmissions_index')
        else:
            return render(request, "customers/index.html", context)

def get_invoices(request):
    customer_id = request.GET.get('customer_id')
    invoices = Invoices.objects.filter(Q(type="ذمم") & Q(require_id__id=customer_id) & Q(tamount__gt=F('customer_paid')))
    invoice_list = list(invoices.values('number'))
    return JsonResponse({'invoices': invoice_list})

def get_invoice_details(request):
    invoice_number = request.GET.get('invoice_number')
    invoice = Invoices.objects.filter(number=invoice_number).first()
    if invoice:
        amount_due = invoice.tamount - invoice.customer_paid
        return JsonResponse({'amount_due': amount_due})
    return JsonResponse({'amount_due': 0})

def customer_details(request, id):
    customer = get_object_or_404(Customers, id=id)
    papers = Papers.objects.filter(customer__id=id)
    vouchers = Vouchers.objects.filter(car_owner__id=id).distinct('car_num', 'car_type')
    invoices = Invoices.objects.filter(Q(require_id__id=id) & Q(type="ذمم")).annotate(
        account=V('Created invoice', output_field=CharField()),
        number_annotated=F('number'),
        amount_annotated=F('tamount'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'created_at_annotated'
    )
    treasury = Treasury.objects.filter(customer__id=id).annotate(
        account=V('Paid to treasury', output_field=CharField()),
        number_annotated=F('number'),
        amount_annotated=F('amount'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'created_at_annotated'
    )
    combined_entries = invoices.union(treasury).order_by('created_at_annotated')
    context = {
        'active_tab': 'customers',
        'id': id,
        'vouchers': vouchers,
        'customer': customer,
        'papers': papers,
        'combined_entries': combined_entries,
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('customers_index')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('customers_index')
    else:
        return render(request, "customers/details.html", context)

def customer_balance(request, id):
    customer = get_object_or_404(Customers, id=id)
    customer_name = customer.name
    invoices = Invoices.objects.filter(Q(require_id__id=id) & Q(type="ذمم"))
    treasury = Treasury.objects.filter(customer__id=id)
    total_invoiced = invoices.aggregate(Sum('tamount'))['tamount__sum'] or 0
    total_treasury = treasury.aggregate(Sum('amount'))['amount__sum'] or 0
    required_amount = total_invoiced - total_treasury
    context = {
        'id': id,
        'customer_name': customer_name,
        'required_amount': required_amount,
        'invoices': invoices,
        'treasury': treasury,
        'active_tab': 'customers',
        'total_invoiced': total_invoiced,
        'total_treasury': total_treasury
    }
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('customers_index')
        else:
            messages.success(request, "There was an error for username or password...")
            return redirect('customers_index')
    else:
        return render(request, "customers/balance.html", context)

@login_required
def add_customer(request):
    if request.method == 'POST':
        name = request.POST.get('customer_name').strip()
        address = request.POST.get('customer_address')
        mail = request.POST.get('customer_mail')
        phone = request.POST.get('customer_phone')
        payment = request.POST.get('customer_payment')
        username = request.POST.get('username')
        password = request.POST.get('password')
        notes = request.POST.get('customer_notes')
        logo = request.FILES.get('customer_logo')
        system_user = request.user.username
        if username and password and mail:
            queryset = Customers.objects.filter(Q(name=name) | Q(phone=phone) | Q(mail=mail) | (Q(username=username) & Q(password=password)))
        elif username and password:
            queryset = Customers.objects.filter(Q(name=name) | Q(phone=phone) | (Q(username=username) & Q(password=password)))
        elif mail:
            queryset = Customers.objects.filter(Q(name=name) | Q(phone=phone) | Q(mail=mail))
        else:
            queryset = Customers.objects.filter(Q(name=name) | Q(phone=phone))
        if queryset.exists():
            record = queryset.first()  # This will get the first object that matches the query
        else:
            record = None
        if record:
            messages.success(request, "Repeated Data...")
        else:
            if logo:
                fs = FileSystemStorage()
                logo_name = fs.save(logo.name, logo)
                # Save to the database
                Customers.objects.create(name=name, mail=mail, phone=phone, address=address, username=username, password=password, notes=notes, payment=payment, logo=logo_name, created_by=system_user)
            else:
                Customers.objects.create(name=name, mail=mail, phone=phone, address=address, username=username, password=password, notes=notes, payment=payment, created_by=system_user)
    return redirect('customers_index')

@login_required
def edit_customer(request, id):
    customer = get_object_or_404(Customers, id=id)
    if request.method == 'POST':
        customer.name = request.POST.get('name').strip()
        customer.address = request.POST.get('address')
        customer.phone = request.POST.get('phone')
        customer.mail = request.POST.get('mail')
        customer.payment = request.POST.get('payment')
        customer.username = request.POST.get('username')
        customer.password = request.POST.get('password')
        customer.notes = request.POST.get('notes')
        customer_logo = request.FILES.get('logo')
        if customer_logo:
            fs = FileSystemStorage()
            logo_name = fs.save(customer_logo.name, customer_logo)
            customer.logo = logo_name
        tcustomer = Customers.objects.exclude(id=id)
        if not tcustomer:
            customer.save()
        else:
            if request.POST.get('username') and request.POST.get('password') and request.POST.get('mail'):
                queryset = tcustomer.filter(Q(name=request.POST.get('name')) | Q(phone=request.POST.get('phone')) | Q(mail=request.POST.get('mail')) | (Q(username=request.POST.get('username')) & Q(password=request.POST.get('password'))))
            elif request.POST.get('username') and request.POST.get('password'):
                queryset = tcustomer.filter(Q(name=request.POST.get('name')) | Q(phone=request.POST.get('phone')) | (Q(username=request.POST.get('username')) & Q(password=request.POST.get('password'))))
            elif request.POST.get('mail'):
                queryset = tcustomer.filter(Q(name=request.POST.get('name')) | Q(phone=request.POST.get('phone')) | Q(mail=request.POST.get('mail')))
            else:
                queryset = tcustomer.filter(Q(name=request.POST.get('name')) | Q(phone=request.POST.get('phone')))
            if queryset.exists():
                record = queryset.first()  # This will get the first object that matches the query
            else:
                record = None
            if record:
                messages.success(request, "Repeated Data...")
            else:
                customer.save()
    return redirect('customers_index')

@login_required
def add_attachment(request):
    if request.method == 'POST':
        customer_id = request.POST.get('customer_id')
        customer = get_object_or_404(Customers, id=customer_id)
        description = request.POST.get('description')
        attachment = request.FILES.get('attachment')
        fs = FileSystemStorage()
        attachment_name = fs.save(attachment.name, attachment)
        # Save to the database
        Papers.objects.create(customer=customer, description=description, paper=attachment_name)
        related_url = reverse('customer_details', kwargs={'id': customer_id})
        return HttpResponseRedirect(related_url)

@login_required
def edit_attachment(request, id):
    paper = get_object_or_404(Papers, id=id)
    if request.method == 'POST':
        customer_id = paper.customer.id
        paper.description = request.POST.get('description')
        customer_paper = request.FILES.get('paper')
        if customer_paper:
            fs = FileSystemStorage()
            paper_name = fs.save(customer_paper.name, customer_paper)
            paper.paper = paper_name
        paper.save()
        related_url = reverse('customer_details', kwargs={'id': customer_id})
        return HttpResponseRedirect(related_url)

@login_required
def delete_attachment(request, id):
    paper = get_object_or_404(Papers, id=id)
    if request.method == 'POST':
        customer_id = paper.customer.id
        paper.delete()
        related_url = reverse('customer_details', kwargs={'id': customer_id})
        return HttpResponseRedirect(related_url)

@login_required
def add_treasury(request):
    if request.method == 'POST':
        customer_id = request.POST['customer_id']
        customer = get_object_or_404(Customers, id=customer_id)
        invoice_number = request.POST['invoice_number']
        invoice = get_object_or_404(Invoices, number=invoice_number)
        number = request.POST['number']
        amount = request.POST['amount']
        payment = request.POST['method']
        # Add the new item to the database
        new_treasury = Treasury(customer=customer, invoice=invoice, number=number, amount=amount, payment=payment)
        new_treasury.save()
        update_customer_balance(customer_id)
        update_invoice(invoice.id)
    return redirect('customers_index')

@login_required
def edit_treasury(request, id):
    if request.method == 'POST':
        customer_id = request.POST['customer_id']
        item = get_object_or_404(Treasury, id=id)
        item.amount = request.POST['edit-amount']
        item.payment = request.POST['method']
        item.save()
        update_customer_balance(customer_id)
        update_invoice(item.invoice.id)
        related_url = reverse('customer_balance', kwargs={'id': customer_id})
        return HttpResponseRedirect(related_url)

@login_required
def delete_treasury(request, id):
    item = get_object_or_404(Treasury, id=id)
    invoice_id = item.invoice.id
    if request.method == 'POST':
        customer_id = request.POST['customer_id']
        item.delete()
        update_customer_balance(customer_id)
        update_invoice(invoice_id)
        related_url = reverse('customer_balance', kwargs={'id': customer_id})
        return HttpResponseRedirect(related_url)

@login_required
def export_treasury(request, id, date_from, date_to):
    customer = get_object_or_404(Customers, id=id)
    customer_name = customer.name
    invoices = Invoices.objects.filter(Q(require_id__id=id) & Q(type="ذمم") & (Q(date__range=(date_from, date_to)) | Q(tamount__gt=F('customer_paid')))).annotate(
        account=V('فاتورة', output_field=CharField()),
        number_annotated=F('number'),
        amount_annotated=F('tamount'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'created_at_annotated'
    )

    treasury = Treasury.objects.filter(Q(customer__id=id) & (Q(invoice__date__range=(date_from, date_to)) | Q(invoice__tamount__gt=F('invoice__customer_paid')))).annotate(
        account=V('سداد', output_field=CharField()),
        number_annotated=F('invoice__number'),
        amount_annotated=F('amount'),
        created_at_annotated=F('created_at')
    ).values(
        'account',
        'number_annotated',
        'amount_annotated',
        'created_at_annotated'
    )
    combined_entries = invoices.union(treasury).order_by('created_at_annotated')
    source_excel_path = os.path.join(settings.BASE_DIR, 'excel_templates', 'Customer Account Sheet.xlsx')
    output_excel_path = f"{os.path.splitext(source_excel_path)[0]}_{customer_name}.xlsx"
    # Load the existing workbook (source file)
    wb = openpyxl.load_workbook(source_excel_path)
    ws = wb.active
    cell = ws.cell(row=2, column=1)
    cell.value = "كشف حركة الحساب الخاص بالعميل/ " + customer_name
    white_fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in range(4, 33428):
        ws.row_dimensions[row].hidden = False
    for row in ws.iter_rows(min_row=4, max_row=33428, max_col=4):
        for cell in row:
            cell.value = None
            cell.fill = white_fill
    idx = start_row = 4
    yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    amman_tz = pytz.timezone('Asia/Amman')
    for idx, item in enumerate(combined_entries, start=start_row):
        item['created_at_annotated'] = localtime(item['created_at_annotated'], amman_tz)
        ws.cell(row=idx, column=1, value=item['account'])
        ws.cell(row=idx, column=2, value=item['number_annotated'])
        ws.cell(row=idx, column=3, value=item['amount_annotated'])
        ws.cell(row=idx, column=4, value=item['created_at_annotated'].replace(tzinfo=None))
        if item['account'] == "سداد":
            for col in range(1, 5):  # Assuming you have data in columns 1 to 4
                ws.cell(row=idx, column=col).fill = yellow_fill
    last_data_row = idx + 1
    for value in range(last_data_row, 33428):
        ws.row_dimensions[value].hidden = True
    wb.save(output_excel_path)
    try:
        output_pdf_path = os.path.splitext(output_excel_path)[0] + '.pdf'
        subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf", output_excel_path, '--outdir', os.path.dirname(output_pdf_path)], check=True)
        zip_path = f"{os.path.splitext(output_excel_path)[0]}.zip"
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(output_excel_path, os.path.basename(output_excel_path))
            zipf.write(output_pdf_path, os.path.basename(output_pdf_path))
        if os.path.exists(zip_path):
            with open(zip_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename="Customer_Report_{id}.zip"'
                return response
        else:
            raise Http404("ZIP file not found")
    except Exception as e:
        print("Error:", e)
        raise Http404("An error occurred while generating the files.")

def filter_invoices(request, id):
    date_from = datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d')
    date_to = datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d')
    invoices = Invoices.objects.filter(Q(require_id__id=id) & Q(type="ذمم") & Q(date__range=(date_from, date_to))).order_by('number')
    amman_tz = pytz.timezone('Asia/Amman')
    invoices_data = []
    for invoice in invoices:
        invoice.created_at = localtime(invoice.created_at, amman_tz)
        invoices_data.append({
            'number': invoice.number,
            'driver__name': invoice.driver.name,
            'tamount': invoice.tamount,
            'created_at': invoice.created_at
        })
    treasury = Treasury.objects.filter(Q(invoice__require_id__id=id) & Q(invoice__type="ذمم") & Q(invoice__date__range=(date_from, date_to))).order_by('created_at')
    treasury_data = []
    for tr in treasury:
        tr.created_at = localtime(tr.created_at, amman_tz)
        treasury_data.append({
            'id': tr.id,
            'number': tr.number,
            'inumber': tr.invoice.number,
            'amount': tr.amount,
            'payment': tr.payment,
            'created_at': tr.created_at
        })
    total_invoiced = invoices.aggregate(Sum('tamount'))['tamount__sum'] or 0
    total_treasury = treasury.aggregate(Sum('amount'))['amount__sum'] or 0
    required_amount = total_invoiced - total_treasury
    return JsonResponse({'invoices': invoices_data, 'treasury': treasury_data, 'total_invoiced': total_invoiced, 'total_treasury': total_treasury, 'required_amount': required_amount}, safe=False)

def update_customer_balance(customer_id):
    customer = get_object_or_404(Customers, id=customer_id)
    total_invoiced = Invoices.objects.filter(Q(require_id__id=customer_id) & Q(type="ذمم")).aggregate(Sum('tamount'))['tamount__sum'] or 0
    total_treasury = Treasury.objects.filter(customer__id=customer_id).aggregate(Sum('amount'))['amount__sum'] or 0
    customer.balance = total_invoiced - total_treasury
    customer.save()

